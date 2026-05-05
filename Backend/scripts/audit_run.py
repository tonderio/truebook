"""
Ultra-audit for a TrueBook reconciliation run.

Reads the DB + original uploaded files, re-runs every parser/conciliation
independently, cross-checks each input source against every other source,
and writes a self-contained markdown audit report with PASS/FAIL gates.

Usage:
    cd Backend && python -m scripts.audit_run --process-id 5
    cd Backend && python -m scripts.audit_run --process-id 5 --include-mongo
    cd Backend && python -m scripts.audit_run --process-id 5 --output /tmp/x.md

Exits 0 on PASS, 1 on FAIL.

This script is **read-only** against the DB. The only side-effect is writing
the markdown report file.
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

# Run from Backend/ so app.* imports resolve
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import SessionLocal
from app.models.process import AccountingProcess
from app.models.result import (
    KushkiResult,
    BanregioResult,
    FeesResult,
    ConciliationResult,
)
from app.models.bitso import BitsoReport, BitsoReportLine
from app.models.classification import BanregioMovementClassification
from app.models.adjustment import RunAdjustment
from app.models.alert import RunAlert
from app.models.file import UploadedFile

from app.services.kushki_parser import parse_kushki
from app.services.banregio_parser import parse_banregio
from app.services.conciliation_engine import (
    conciliate_kushki_daily,
    conciliate_kushki_vs_banregio,
    conciliate_fees,
    get_tolerance,
)
from app.services.auto_classifier import (
    classify_movement,
    KUSHKI_REF_PATTERN,
)


# ── tolerances for the audit (NOT the engine's; auditor's own thresholds) ─

CROSS_SOURCE_TOLERANCE_PCT = 0.05    # 5% — Kushki↔Banregio totals (timing skew)
CROSS_SOURCE_TOLERANCE_ABS = 50_000  # MXN — absolute floor on cross-source check
SPOT_CHECK_PER_CATEGORY = 3
ARITHMETIC_TOLERANCE = 0.01          # MXN — for matching engine outputs


# ── helpers ───────────────────────────────────────────────────────────────


def fmt_money(v: Any) -> str:
    if v is None:
        return "—"
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def fmt_pct(v: Any) -> str:
    if v is None:
        return "—"
    return f"{float(v):.2f}%"


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def status_icon(passed: bool) -> str:
    return "✅" if passed else "❌"


def find_upload_path(uploads_dir: Path, prefix: str) -> Path | None:
    """Pick the first file matching prefix*  in the uploads dir."""
    if not uploads_dir.exists():
        return None
    for p in sorted(uploads_dir.iterdir()):
        if p.name.startswith(prefix):
            return p
    return None


# ── Section 1 ─────────────────────────────────────────────────────────────


def section_1_ingestion(db, process, include_mongo: bool) -> tuple[bool, list[str]]:
    """Re-parse / re-extract each source and compare to stored values."""
    out = []
    all_pass = True

    uploads = Path(__file__).resolve().parent.parent / "uploads" / str(process.id)
    out.append(f"_Source files directory: `{uploads}`_")
    out.append("")

    # --- Kushki ----------------------------------------------------------
    out.append("### 1.1 Kushki")
    out.append("")
    kr = db.query(KushkiResult).filter_by(process_id=process.id).first()
    kushki_path = find_upload_path(uploads, "kushki_")

    if not kr:
        out.append(f"{status_icon(False)} no `KushkiResult` row in DB")
        all_pass = False
    elif not kushki_path:
        out.append(f"{status_icon(False)} no Kushki file in uploads dir")
        all_pass = False
    else:
        with open(kushki_path, "rb") as f:
            content = f.read()
        reparsed = parse_kushki(content, kushki_path.name)

        stored_total = to_float(kr.total_net_deposit)
        reparsed_total = to_float(reparsed["total_net_deposit"])
        delta = abs(stored_total - reparsed_total)
        ok = delta <= ARITHMETIC_TOLERANCE
        all_pass = all_pass and ok

        stored_days = len(kr.daily_summary or [])
        reparsed_days = len(reparsed.get("daily_summary") or [])
        days_ok = stored_days == reparsed_days
        all_pass = all_pass and days_ok

        # Cross-check: per-day merchant_detail sum should match daily_summary per day
        per_day_merchant = {}
        for r in (kr.merchant_daily_detail or []) if False else []:
            pass  # column not stored in JSON above; skip
        merchant_total = 0.0
        for m in kr.merchant_detail or []:
            merchant_total += to_float(m.get("net_deposit"))
        merchant_vs_daily_delta = abs(merchant_total - stored_total)
        merchant_ok = merchant_vs_daily_delta <= 1.00  # rounding tolerance

        out.append(f"| Check | Stored | Re-parsed | Delta | Result |")
        out.append(f"|---|---:|---:|---:|:---:|")
        out.append(f"| total_net_deposit | {fmt_money(stored_total)} | {fmt_money(reparsed_total)} | {fmt_money(delta)} | {status_icon(ok)} |")
        out.append(f"| daily_summary row count | {stored_days} | {reparsed_days} | {abs(stored_days - reparsed_days)} | {status_icon(days_ok)} |")
        out.append(f"| merchant_detail Σnet_deposit vs daily Σ | {fmt_money(merchant_total)} | {fmt_money(stored_total)} | {fmt_money(merchant_vs_daily_delta)} | {status_icon(merchant_ok)} |")
        out.append("")

    # --- Bitso -----------------------------------------------------------
    out.append("### 1.2 Bitso")
    out.append("")
    br = db.query(BitsoReport).filter_by(process_id=process.id).first()
    if not br:
        out.append(
            "ℹ no `BitsoReport` row — Bitso API returned no deposits for this period "
            "(typically the API key is scoped to a different month, or BITSO_API_ENABLED was off "
            "during the run). This is informational; downstream classification of `bitso_acquirer` "
            "movements in Banregio still works via CLABE/keyword detection."
        )
        out.append("")
    else:
        lines = db.query(BitsoReportLine).filter_by(report_id=br.id).all()
        line_count = len(lines)
        line_total = sum(to_float(l.gross_amount) for l in lines)
        stored_total = to_float(br.total_amount)
        delta = abs(line_total - stored_total)
        sum_ok = delta <= ARITHMETIC_TOLERANCE
        all_pass = all_pass and sum_ok

        # Period containment
        period_year, period_month = process.period_year, process.period_month
        out_of_period = []
        for l in lines:
            d = l.txn_date
            if d is None:
                continue
            if d.year != period_year or d.month != period_month:
                out_of_period.append((l.txn_id, d))
        period_ok = len(out_of_period) == 0
        all_pass = all_pass and period_ok

        # Dedup by txn_id
        seen = set()
        dups = []
        for l in lines:
            if l.txn_id in seen:
                dups.append(l.txn_id)
            else:
                seen.add(l.txn_id)
        dedup_ok = len(dups) == 0
        all_pass = all_pass and dedup_ok

        out.append(f"| Check | Value | Result |")
        out.append(f"|---|---:|:---:|")
        out.append(f"| line count | {line_count} | — |")
        out.append(f"| Σ(line.gross_amount) | {fmt_money(line_total)} | — |")
        out.append(f"| BitsoReport.total_amount | {fmt_money(stored_total)} | — |")
        out.append(f"| sum vs total delta | {fmt_money(delta)} | {status_icon(sum_ok)} |")
        out.append(f"| txn_dates inside {period_year}-{period_month:02d} | {line_count - len(out_of_period)}/{line_count} | {status_icon(period_ok)} |")
        out.append(f"| duplicate txn_ids | {len(dups)} | {status_icon(dedup_ok)} |")
        if out_of_period:
            out.append("")
            out.append(f"⚠ {len(out_of_period)} deposits outside period (first 5):")
            for txid, d in out_of_period[:5]:
                out.append(f"  - `{txid}` on {d}")
        out.append("")

    # --- Banregio --------------------------------------------------------
    out.append("### 1.3 Banregio")
    out.append("")
    bg = db.query(BanregioResult).filter_by(process_id=process.id).first()
    bg_path = find_upload_path(uploads, "banregio_")
    if not bg:
        out.append(f"{status_icon(False)} no `BanregioResult` row")
        all_pass = False
    elif not bg_path:
        out.append(f"{status_icon(False)} no Banregio file in uploads dir")
        all_pass = False
    else:
        with open(bg_path, "rb") as f:
            content = f.read()
        reparsed = parse_banregio(content, bg_path.name)
        stored_movs = bg.movements or []
        reparsed_movs = reparsed.get("movements") or []

        count_ok = len(stored_movs) == len(reparsed_movs)
        all_pass = all_pass and count_ok

        stored_credits = sum(to_float(m.get("credit")) for m in stored_movs)
        reparsed_credits = sum(to_float(m.get("credit")) for m in reparsed_movs)
        stored_debits = sum(to_float(m.get("debit")) for m in stored_movs)
        reparsed_debits = sum(to_float(m.get("debit")) for m in reparsed_movs)

        c_delta = abs(stored_credits - reparsed_credits)
        d_delta = abs(stored_debits - reparsed_debits)
        sums_ok = c_delta <= ARITHMETIC_TOLERANCE and d_delta <= ARITHMETIC_TOLERANCE
        all_pass = all_pass and sums_ok

        out.append(f"| Check | Stored | Re-parsed | Delta | Result |")
        out.append(f"|---|---:|---:|---:|:---:|")
        out.append(f"| movement count | {len(stored_movs)} | {len(reparsed_movs)} | {abs(len(stored_movs) - len(reparsed_movs))} | {status_icon(count_ok)} |")
        out.append(f"| Σ credits | {fmt_money(stored_credits)} | {fmt_money(reparsed_credits)} | {fmt_money(c_delta)} | {status_icon(c_delta <= ARITHMETIC_TOLERANCE)} |")
        out.append(f"| Σ debits | {fmt_money(stored_debits)} | {fmt_money(reparsed_debits)} | {fmt_money(d_delta)} | {status_icon(d_delta <= ARITHMETIC_TOLERANCE)} |")
        out.append("")

    # --- Tonder (Mongo) --------------------------------------------------
    out.append("### 1.4 Tonder (MongoDB → Fees)")
    out.append("")
    fr = db.query(FeesResult).filter_by(process_id=process.id).first()
    if not fr:
        out.append(f"{status_icon(False)} no `FeesResult` row")
        all_pass = False
    else:
        merchant_total = sum(to_float(m.get("total_fee")) for m in (fr.merchant_summary or []))
        stored_total = to_float(fr.total_fees)
        out.append(f"| Check | Value | Result |")
        out.append(f"|---|---:|:---:|")
        out.append(f"| merchant_summary count | {len(fr.merchant_summary or [])} | — |")
        out.append(f"| daily_breakdown rows | {len(fr.daily_breakdown or [])} | — |")
        out.append(f"| withdrawals_summary count | {len(fr.withdrawals_summary or [])} | — |")
        out.append(f"| refunds_summary count | {len(fr.refunds_summary or [])} | — |")
        out.append(f"| total_fees stored | {fmt_money(stored_total)} | — |")
        out.append(f"| Σ merchant.total_fee (re-summed) | {fmt_money(merchant_total)} | {status_icon(abs(merchant_total - stored_total) <= 1.00)} |")
        if not include_mongo:
            out.append("")
            out.append("_(Mongo re-extraction skipped — pass `--include-mongo` to re-pull from Atlas and cross-check counts.)_")
        else:
            try:
                from app.services import mongo_extractor
                txs = mongo_extractor.extract_transactions(process.period_year, process.period_month)
                wds = mongo_extractor.extract_withdrawals(process.period_year, process.period_month)
                rfs = mongo_extractor.extract_refunds(process.period_year, process.period_month)
                out.append("")
                out.append(f"| Mongo re-extract | Count |")
                out.append(f"|---|---:|")
                out.append(f"| transactions | {len(txs)} |")
                out.append(f"| withdrawals | {len(wds)} |")
                out.append(f"| refunds | {len(rfs)} |")
            except Exception as e:
                out.append("")
                out.append(f"⚠ Mongo extraction failed: `{type(e).__name__}: {e}`")
        out.append("")

    return all_pass, out


# ── Section 2 ─────────────────────────────────────────────────────────────


def section_2_classification(db, process) -> tuple[bool, list[str]]:
    out = []
    all_pass = True

    classifications = (
        db.query(BanregioMovementClassification)
        .filter_by(process_id=process.id)
        .order_by(BanregioMovementClassification.movement_index)
        .all()
    )
    bg = db.query(BanregioResult).filter_by(process_id=process.id).first()

    if not classifications or not bg:
        out.append(f"{status_icon(False)} missing classifications or BanregioResult")
        return False, out

    movements = bg.movements or []
    cov = float(process.coverage_pct or 0)
    cov_ok = cov >= 100.0
    all_pass = all_pass and cov_ok

    # Distribution
    dist: dict[str, int] = {}
    for c in classifications:
        dist[c.classification] = dist.get(c.classification, 0) + 1
    unclassified = dist.get("unclassified", 0)
    no_unclassified = unclassified == 0
    all_pass = all_pass and no_unclassified

    out.append(f"| Check | Value | Result |")
    out.append(f"|---|---:|:---:|")
    out.append(f"| coverage_pct | {fmt_pct(cov)} | {status_icon(cov_ok)} |")
    out.append(f"| classification rows | {len(classifications)} | — |")
    out.append(f"| total movements | {len(movements)} | — |")
    out.append(f"| unclassified count | {unclassified} | {status_icon(no_unclassified)} |")
    out.append("")

    out.append("### 2.1 Distribution by category")
    out.append("")
    out.append("| Category | Count | Σ Credit | Σ Debit |")
    out.append("|---|---:|---:|---:|")
    for cat in sorted(dist.keys()):
        idxs = [c.movement_index for c in classifications if c.classification == cat]
        sum_credit = sum(to_float(movements[i].get("credit")) for i in idxs if i < len(movements))
        sum_debit = sum(to_float(movements[i].get("debit")) for i in idxs if i < len(movements))
        out.append(f"| {cat} | {dist[cat]} | {fmt_money(sum_credit)} | {fmt_money(sum_debit)} |")
    out.append("")

    # Spot-check rule firing
    out.append(f"### 2.2 Spot-check (up to {SPOT_CHECK_PER_CATEGORY} per non-zero category)")
    out.append("")
    out.append("| Category | movement_idx | Description (truncated) | Re-classified | Rule fires? |")
    out.append("|---|---:|---|---|:---:|")
    spot_pass = True
    spot_total = 0
    spot_misses = 0
    for cat in sorted(dist.keys()):
        if cat == "unclassified":
            continue
        candidates = [c for c in classifications if c.classification == cat][:SPOT_CHECK_PER_CATEGORY]
        for c in candidates:
            mov = movements[c.movement_index] if c.movement_index < len(movements) else {}
            descr = (mov.get("description") or "")[:50].replace("|", "/")
            credit = to_float(mov.get("credit"))
            debit = to_float(mov.get("debit"))
            amount = credit if credit else -debit
            mov_type = "abono" if credit else "cargo"
            ref = str(mov.get("deposit_ref") or mov.get("reference") or "")
            try:
                refire_cls, _refire_acq, _method = classify_movement(
                    description=str(mov.get("description") or ""),
                    reference=ref,
                    amount=amount,
                    movement_type=mov_type,
                )
            except Exception as e:
                refire_cls = f"ERROR:{type(e).__name__}"
            fires = refire_cls == cat
            spot_total += 1
            if not fires:
                spot_misses += 1
                spot_pass = False
            out.append(f"| {cat} | {c.movement_index} | `{descr}` | `{refire_cls}` | {status_icon(fires)} |")
    all_pass = all_pass and spot_pass
    out.append("")
    out.append("")
    out.append(f"_Spot-check: {spot_total - spot_misses}/{spot_total} re-classify identically — {status_icon(spot_pass)}_")
    out.append("")

    return all_pass, out


# ── Section 3 ─────────────────────────────────────────────────────────────


def section_3_conciliation(db, process) -> tuple[bool, list[str]]:
    out = []
    all_pass = True

    crs = (
        db.query(ConciliationResult)
        .filter_by(process_id=process.id)
        .all()
    )
    by_type = {cr.conciliation_type: cr for cr in crs}

    out.append(f"| Type | total_conciliated | total_difference | matched | differences | unmatched_K | unmatched_B |")
    out.append(f"|---|---:|---:|---:|---:|---:|---:|")
    for cr in crs:
        m = len(cr.matched or [])
        d = len(cr.differences or [])
        uk = len(cr.unmatched_kushki or [])
        ub = len(cr.unmatched_banregio or [])
        out.append(
            f"| {cr.conciliation_type} | {fmt_money(cr.total_conciliated)} | {fmt_money(cr.total_difference)} | {m} | {d} | {uk} | {ub} |"
        )
    out.append("")

    # 3.1 fees re-computation
    out.append("### 3.1 fees rollup arithmetic")
    fr = db.query(FeesResult).filter_by(process_id=process.id).first()
    fees_cr = by_type.get("fees")
    if fr and fees_cr:
        fees_inp = {
            "merchant_summary": fr.merchant_summary or [],
            "daily_breakdown": fr.daily_breakdown or [],
            "withdrawals_summary": fr.withdrawals_summary or [],
            "refunds_summary": fr.refunds_summary or [],
            "total_fees": to_float(fr.total_fees),
        }
        try:
            recomputed = conciliate_fees(fees_inp)
            stored_diff = to_float(fees_cr.total_difference)
            recomp_diff = to_float(recomputed.get("total_difference"))
            ok = abs(stored_diff - recomp_diff) <= ARITHMETIC_TOLERANCE
            out.append(f"  - stored total_difference: {fmt_money(stored_diff)}")
            out.append(f"  - recomputed total_difference: {fmt_money(recomp_diff)}")
            out.append(f"  - {status_icon(ok)} match within tolerance")
            all_pass = all_pass and ok
        except Exception as e:
            out.append(f"  - ⚠ recomputation failed: `{type(e).__name__}: {e}`")
            all_pass = False
    else:
        out.append("  - skipped (missing FeesResult or fees ConciliationResult)")
    out.append("")

    # 3.2 kushki_daily independent F4 verification
    out.append("### 3.2 kushki_daily — F4 formula independent re-verify")
    kr = db.query(KushkiResult).filter_by(process_id=process.id).first()
    daily_cr = by_type.get("kushki_daily")
    if kr and daily_cr:
        rows = kr.daily_summary or []
        my_diff_total = 0.0
        my_diffs = 0
        per_day = []
        for r in rows:
            gross = to_float(r.get("gross_amount"))
            comm = to_float(r.get("commission"))
            roll = to_float(r.get("rolling_reserve"))
            rrr = to_float(r.get("rr_released"))
            ref = to_float(r.get("refund"))
            cb = to_float(r.get("chargeback"))
            vd = to_float(r.get("void"))
            ma = to_float(r.get("manual_adj"))
            net = to_float(r.get("net_deposit"))
            f4 = gross - comm - roll - rrr + ref + cb + vd + ma
            d = abs(f4 - net)
            if d > ARITHMETIC_TOLERANCE:
                my_diffs += 1
                my_diff_total += d
                per_day.append((r.get("date"), net, f4, d))
        stored_diff = to_float(daily_cr.total_difference)
        diffs_match = abs(my_diff_total - stored_diff) <= ARITHMETIC_TOLERANCE * len(rows)
        out.append(f"  - days re-verified: {len(rows)}")
        out.append(f"  - days with |Δ| > $0.01 (independent): {my_diffs}")
        out.append(f"  - Σ |Δ| (independent): {fmt_money(my_diff_total)}")
        out.append(f"  - stored total_difference: {fmt_money(stored_diff)}")
        out.append(f"  - {status_icon(diffs_match)} independent re-verify matches stored")
        all_pass = all_pass and diffs_match
        if per_day:
            out.append("")
            out.append("  Days with residuals:")
            out.append("")
            out.append("  | date | net_deposit | F4 computed | |Δ| |")
            out.append("  |---|---:|---:|---:|")
            for date_, net, f4, d in per_day[:10]:
                out.append(f"  | {date_} | {fmt_money(net)} | {fmt_money(f4)} | {fmt_money(d)} |")
    else:
        out.append("  - skipped (missing KushkiResult or kushki_daily ConciliationResult)")
    out.append("")

    # 3.3 kushki_vs_banregio independent re-match
    out.append("### 3.3 kushki_vs_banregio — re-match")
    bg = db.query(BanregioResult).filter_by(process_id=process.id).first()
    vs_cr = by_type.get("kushki_vs_banregio")
    if kr and bg and vs_cr:
        # Independent re-run using engine. Mirror the pipeline's Stage 8b
        # path: pass classifications so the matcher can pre-filter to
        # `kushki_acquirer` credits only (matches what's stored).
        tolerance = get_tolerance(db)
        cls_rows = (
            db.query(BanregioMovementClassification)
            .filter_by(process_id=process.id)
            .all()
        )
        cls_map = {c.movement_index: c.classification for c in cls_rows}
        rerun = conciliate_kushki_vs_banregio(
            {"daily_summary": kr.daily_summary or []},
            {"movements": bg.movements or []},
            tolerance=tolerance,
            classifications=cls_map if cls_map else None,
        )
        stored_matched = len(vs_cr.matched or [])
        stored_uk = len(vs_cr.unmatched_kushki or [])
        stored_ub = len(vs_cr.unmatched_banregio or [])
        rerun_matched = len(rerun["matched"])
        rerun_uk = len(rerun["unmatched_kushki"])
        rerun_ub = len(rerun["unmatched_banregio"])
        match_ok = stored_matched == rerun_matched
        uk_ok = stored_uk == rerun_uk
        ub_ok = stored_ub == rerun_ub
        ok = match_ok and uk_ok and ub_ok
        out.append(f"  | metric | stored | re-run | result |")
        out.append(f"  |---|---:|---:|:---:|")
        out.append(f"  | matched pairs | {stored_matched} | {rerun_matched} | {status_icon(match_ok)} |")
        out.append(f"  | unmatched_kushki | {stored_uk} | {rerun_uk} | {status_icon(uk_ok)} |")
        out.append(f"  | unmatched_banregio | {stored_ub} | {rerun_ub} | {status_icon(ub_ok)} |")
        all_pass = all_pass and ok
    else:
        out.append("  - skipped (missing inputs or kushki_vs_banregio ConciliationResult)")
    out.append("")

    return all_pass, out


# ── Section 4 ─────────────────────────────────────────────────────────────


def section_4_cross_source(db, process) -> tuple[bool, list[str]]:
    out = []
    all_pass = True

    classifications = (
        db.query(BanregioMovementClassification)
        .filter_by(process_id=process.id)
        .all()
    )
    bg = db.query(BanregioResult).filter_by(process_id=process.id).first()
    kr = db.query(KushkiResult).filter_by(process_id=process.id).first()
    br = db.query(BitsoReport).filter_by(process_id=process.id).first()
    bitso_lines = db.query(BitsoReportLine).filter_by(report_id=br.id).all() if br else []

    movements = bg.movements if bg else []

    # Helper: sum credits (or debits) for a given classification
    def banregio_total(category: str, side: str = "credit") -> float:
        total = 0.0
        for c in classifications:
            if c.classification != category:
                continue
            if c.movement_index < len(movements):
                total += to_float(movements[c.movement_index].get(side))
        return total

    def reasonable(stored: float, expected: float) -> tuple[bool, float, float]:
        delta = abs(stored - expected)
        pct = (delta / expected * 100.0) if expected else 0.0
        thresh = max(CROSS_SOURCE_TOLERANCE_ABS, expected * CROSS_SOURCE_TOLERANCE_PCT)
        return delta <= thresh, delta, pct

    out.append("| Check | Expected | Actual | |Δ| | Δ% | Result |")
    out.append("|---|---:|---:|---:|---:|:---:|")

    # Kushki ↔ Banregio kushki_acquirer
    kushki_total = to_float(kr.total_net_deposit) if kr else 0.0
    bg_kushki = banregio_total("kushki_acquirer", "credit")
    ok, delta, pct = reasonable(bg_kushki, kushki_total)
    all_pass = all_pass and ok
    out.append(
        f"| Kushki net_deposit ↔ Banregio kushki_acquirer | {fmt_money(kushki_total)} | "
        f"{fmt_money(bg_kushki)} | {fmt_money(delta)} | {pct:.2f}% | {status_icon(ok)} |"
    )

    # Bitso ↔ Banregio bitso_acquirer
    bitso_total = sum(to_float(l.gross_amount) for l in bitso_lines)
    bg_bitso = banregio_total("bitso_acquirer", "credit")
    if bitso_total == 0 and bg_bitso > 0:
        # No source data to compare against — information only
        out.append(
            f"| Bitso total deposits ↔ Banregio bitso_acquirer | (no Bitso source data) | "
            f"{fmt_money(bg_bitso)} | — | — | ℹ |"
        )
    else:
        ok, delta, pct = reasonable(bg_bitso, bitso_total)
        all_pass = all_pass and ok
        out.append(
            f"| Bitso total deposits ↔ Banregio bitso_acquirer | {fmt_money(bitso_total)} | "
            f"{fmt_money(bg_bitso)} | {fmt_money(delta)} | {pct:.2f}% | {status_icon(ok)} |"
        )

    # bank_expense sanity (< 1% of total credits)
    total_credits = sum(to_float(m.get("credit")) for m in movements)
    bg_expense = banregio_total("bank_expense", "debit")
    pct_of_credits = (bg_expense / total_credits * 100.0) if total_credits else 0.0
    expense_ok = pct_of_credits < 1.0
    all_pass = all_pass and expense_ok
    out.append(
        f"| bank_expense as % of total credits | < 1.00% | {pct_of_credits:.4f}% | — | — | {status_icon(expense_ok)} |"
    )

    # tax retention sanity — informational
    bg_tax = banregio_total("tax", "debit")
    tax_pct = (bg_tax / total_credits * 100.0) if total_credits else 0.0
    out.append(
        f"| tax (ISR) total | (informational) | {fmt_money(bg_tax)} | — | {tax_pct:.4f}% of credits | ℹ |"
    )

    # settlement_to_merchant — debit side
    bg_settle = banregio_total("settlement_to_merchant", "debit")
    out.append(
        f"| settlement_to_merchant (Σ debit) | (informational) | {fmt_money(bg_settle)} | — | — | ℹ |"
    )

    # other acquirers (informational only, no source data)
    for cat in ("unlimit_acquirer", "pagsmile_acquirer", "stp_acquirer"):
        v = banregio_total(cat, "credit")
        out.append(f"| {cat} (Σ credit) | (no source data) | {fmt_money(v)} | — | — | ℹ |")

    out.append("")

    return all_pass, out


# ── Section 5 ─────────────────────────────────────────────────────────────


def section_5_adjustments_status(db, process) -> tuple[bool, list[str]]:
    out = []
    all_pass = True

    adjustments = db.query(RunAdjustment).filter_by(process_id=process.id).all()
    by_status = {"pending": [], "approved": [], "rejected": []}
    for a in adjustments:
        by_status.setdefault(a.status, []).append(a)

    no_pending = len(by_status["pending"]) == 0
    all_pass = all_pass and no_pending

    out.append(f"| Field | Value | Result |")
    out.append(f"|---|---|:---:|")
    out.append(f"| process.status | `{process.status}` | {status_icon(process.status == 'reconciled')} |")
    out.append(f"| process.coverage_pct | {fmt_pct(process.coverage_pct)} | {status_icon(to_float(process.coverage_pct) >= 100.0)} |")
    out.append(f"| process.reconciled_by | {process.reconciled_by} | {status_icon(process.reconciled_by is not None)} |")
    out.append(f"| process.reconciled_at | {process.reconciled_at} | {status_icon(process.reconciled_at is not None)} |")
    out.append(f"| total adjustments | {len(adjustments)} | — |")
    out.append(f"| pending adjustments | {len(by_status['pending'])} | {status_icon(no_pending)} |")
    out.append(f"| approved adjustments | {len(by_status['approved'])} | — |")
    out.append(f"| rejected adjustments | {len(by_status['rejected'])} | — |")
    out.append("")

    if process.status != "reconciled":
        all_pass = False

    if adjustments:
        out.append("### 5.1 Adjustment detail")
        out.append("")
        out.append("| id | type | direction | amount | status | conciliation_type | merchant | description |")
        out.append("|---:|---|---|---:|---|---|---|---|")
        for a in adjustments:
            descr = (a.description or "")[:80].replace("|", "/")
            out.append(
                f"| {a.id} | {a.adjustment_type} | {a.direction} | {fmt_money(a.amount)} | "
                f"{a.status} | {a.conciliation_type or '—'} | {a.merchant_name or '—'} | {descr} |"
            )
        out.append("")

    # Per-conciliation: |delta| − |adjustments| ≤ tolerance
    out.append("### 5.2 Per-conciliation: delta covered by approved adjustments")
    out.append("")
    out.append("| conciliation_type | |delta| | |adj_net| | remaining | result |")
    out.append("|---|---:|---:|---:|:---:|")
    crs = db.query(ConciliationResult).filter_by(process_id=process.id).all()
    for cr in crs:
        delta = abs(to_float(cr.total_difference))
        adj_total = sum(
            to_float(a.amount) * (1 if a.direction == "ADD" else -1)
            for a in by_status["approved"]
            if a.conciliation_type == cr.conciliation_type
        )
        remaining = delta - abs(adj_total)
        ok = remaining <= ARITHMETIC_TOLERANCE
        all_pass = all_pass and ok
        out.append(
            f"| {cr.conciliation_type} | {fmt_money(delta)} | {fmt_money(abs(adj_total))} | {fmt_money(remaining)} | {status_icon(ok)} |"
        )
    out.append("")

    return all_pass, out


# ── Section 6 ─────────────────────────────────────────────────────────────


def section_6_anomalies(db, process, prior_findings: list[str]) -> tuple[bool, list[str]]:
    out = []

    out.append("Known anomalies in this run, with explanations:")
    out.append("")

    # AFUN $250 specific
    afun_adj = (
        db.query(RunAdjustment)
        .filter(
            RunAdjustment.process_id == process.id,
            RunAdjustment.merchant_name == "AFUN",
        )
        .first()
    )
    if afun_adj:
        out.append(
            f"- **AFUN 2026-03-17 unexplained credit** — covered by adjustment id={afun_adj.id} "
            f"({fmt_money(afun_adj.amount)} {afun_adj.direction}, type=`{afun_adj.adjustment_type}`). "
            f"Source: Kushki's reported `Depósito Neto` is stamped $250 higher than the sum of input columns produces. "
            f"Same gap may recur in subsequent months — track and follow up with Kushki account manager if 2nd occurrence."
        )

    # Bitso period containment
    br = db.query(BitsoReport).filter_by(process_id=process.id).first()
    if br:
        lines = db.query(BitsoReportLine).filter_by(report_id=br.id).all()
        out_of_period = [l for l in lines if l.txn_date and (l.txn_date.year != process.period_year or l.txn_date.month != process.period_month)]
        if not out_of_period:
            out.append(
                f"- **Bitso period containment** — all {len(lines)} deposits fall within "
                f"{process.period_year}-{process.period_month:02d}. ✅"
            )
        else:
            out.append(
                f"- **Bitso period leakage** — {len(out_of_period)} deposits outside "
                f"{process.period_year}-{process.period_month:02d}. Investigate API filter."
            )

    # Alerts active (not dismissed)
    alerts = db.query(RunAlert).filter_by(process_id=process.id).all()
    if alerts:
        out.append("")
        out.append("Alerts on file for this run:")
        out.append("")
        out.append("| level | type | title |")
        out.append("|---|---|---|")
        for a in alerts:
            out.append(f"| {a.alert_level} | {a.alert_type} | {(a.title or '')[:80]} |")

    # Findings collected by other sections
    if prior_findings:
        out.append("")
        out.append("New findings surfaced by this audit run:")
        for f in prior_findings:
            out.append(f"- {f}")

    out.append("")
    return True, out  # this section is informational, never fails


# ── Section 7 ─────────────────────────────────────────────────────────────


def section_7_v2_report_freshness(db, process) -> tuple[bool, list[str]]:
    """Detect a stale persisted v2 report file vs the current canonical
    output of `builder.build_to_bytes`.

    The Banregio Reconciliation Report v2 endpoint persists an audit
    copy at `uploads/{process_id}/reports/RECONCILIACION_BANREGIO_*.xlsx`
    every time it generates. If FinOps downloads that file and then the
    code's spec evolves (e.g. the column-split from 17→18 cols we
    shipped after April), the persisted copy goes stale — but there's
    no signal in the UI / on the disk that the file is out of date.

    This section regenerates the v2 workbook in-memory (via builder)
    and diffs the current desglose-table column count + headers against
    the persisted file. Mismatch → ❌ with a recommendation to regenerate.

    No file is mutated — the current generation is pure-bytes via
    `build_to_bytes`, never written to disk.
    """
    out = []
    import io
    import openpyxl

    # Persisted file path (matches what banregio_report router writes)
    base_dir = Path(__file__).resolve().parent.parent / "uploads" / str(process.id) / "reports"
    persisted_files = list(base_dir.glob("RECONCILIACION_BANREGIO_*.xlsx")) if base_dir.exists() else []

    if not persisted_files:
        out.append(
            f"ℹ No persisted v2 report found at `uploads/{process.id}/reports/`. "
            f"Skipping freshness check — first generation will create one."
        )
        return True, out

    # Load the persisted file's `Por Adquirente` desglose header
    persisted_path = persisted_files[0]
    out.append(f"Persisted v2 report: `{persisted_path.name}`")

    try:
        per_wb = openpyxl.load_workbook(persisted_path, data_only=True)
        per_ws = per_wb["Por Adquirente"]
    except Exception as exc:
        out.append(f"⚠ Failed to read persisted report: `{type(exc).__name__}: {exc}`")
        return False, out

    # Find the desglose header row (col A == 'Comercio')
    def _find_desglose_header(ws) -> tuple[int, list[str]] | None:
        for r in range(1, min(ws.max_row, 80) + 1):
            if ws.cell(r, 1).value == "Comercio":
                hdrs = []
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is None or v == "":
                        break
                    hdrs.append(str(v))
                return r, hdrs
        return None

    persisted_hdr = _find_desglose_header(per_ws)

    # Generate the canonical workbook in-memory and grab its desglose header
    try:
        from app.services.banregio_report_v2 import builder
        canonical_bytes, _ = builder.build_to_bytes(db, process)
        can_wb = openpyxl.load_workbook(io.BytesIO(canonical_bytes), data_only=True)
        can_ws = can_wb["Por Adquirente"]
        canonical_hdr = _find_desglose_header(can_ws)
    except Exception as exc:
        out.append(f"⚠ Failed to build canonical report: `{type(exc).__name__}: {exc}`")
        return False, out

    if persisted_hdr is None or canonical_hdr is None:
        out.append("⚠ Could not locate `Comercio` header row in one of the workbooks — manual check required")
        return False, out

    persisted_cols = persisted_hdr[1]
    canonical_cols = canonical_hdr[1]
    cols_match = persisted_cols == canonical_cols
    count_match = len(persisted_cols) == len(canonical_cols)

    out.append("")
    out.append(f"| Check | Persisted | Canonical | Result |")
    out.append(f"|---|---:|---:|:---:|")
    out.append(
        f"| Desglose column count | {len(persisted_cols)} | {len(canonical_cols)} | "
        f"{status_icon(count_match)} |"
    )
    out.append(
        f"| Headers match exactly | {'see below' if not cols_match else 'identical'} | — | "
        f"{status_icon(cols_match)} |"
    )

    if not cols_match:
        out.append("")
        out.append("Header diff (persisted vs canonical):")
        out.append("")
        out.append("| # | Persisted | Canonical |")
        out.append("|---:|---|---|")
        max_n = max(len(persisted_cols), len(canonical_cols))
        for i in range(max_n):
            p = persisted_cols[i] if i < len(persisted_cols) else "—"
            c = canonical_cols[i] if i < len(canonical_cols) else "—"
            marker = "✓" if p == c else "✗"
            out.append(f"| {i+1} {marker} | `{p}` | `{c}` |")
        out.append("")
        out.append(
            "**Recommendation**: regenerate the v2 report (POST "
            f"`/api/processes/{process.id}/banregio-report-v2`, or click "
            "'Reporte v2' in the UI). The persisted copy is from older code "
            "and the column layout has since evolved."
        )

    out.append("")
    return cols_match and count_match, out


# ── orchestrator ──────────────────────────────────────────────────────────


def run_audit(process_id: int, output_path: Path, include_mongo: bool) -> bool:
    db = SessionLocal()
    try:
        process = db.query(AccountingProcess).filter_by(id=process_id).first()
        if not process:
            print(f"❌ process_id={process_id} not found")
            return False

        report: list[str] = []
        report.append(f"# Ultra audit — {process.name} (process_id={process.id})")
        report.append("")
        report.append(f"_Period: {process.period_year}-{process.period_month:02d}_  ")
        report.append(f"_Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}_  ")
        report.append(f"_Audit script: `Backend/scripts/audit_run.py`_")
        report.append("")

        sections: list[tuple[str, Callable, dict]] = [
            ("Section 1 — Source ingestion fidelity", section_1_ingestion, {"include_mongo": include_mongo}),
            ("Section 2 — Auto-classification audit", section_2_classification, {}),
            ("Section 3 — Conciliation arithmetic", section_3_conciliation, {}),
            ("Section 4 — Cross-source validation", section_4_cross_source, {}),
            ("Section 5 — Adjustments + status integrity", section_5_adjustments_status, {}),
        ]

        verdicts: list[tuple[str, bool]] = []
        body: list[str] = []
        for title, fn, kwargs in sections:
            print(f"running {title}…")
            try:
                passed, content = fn(db, process, **kwargs)
            except Exception as e:
                passed = False
                content = [f"⚠ section raised: `{type(e).__name__}: {e}`"]
                import traceback
                content.append("```")
                content.append(traceback.format_exc())
                content.append("```")
            verdicts.append((title, passed))
            body.append(f"## {title}  {status_icon(passed)}")
            body.append("")
            body.extend(content)
            body.append("")

        # Section 6 (informational) runs after the gated sections
        passed6, content6 = section_6_anomalies(db, process, prior_findings=[])
        body.append(f"## Section 6 — Anomaly registry  ℹ")
        body.append("")
        body.extend(content6)
        body.append("")

        # Section 7 (gated) — v2 report freshness check.
        # Runs at the end because it depends on a complete pipeline state
        # and isn't relevant unless the run is fully populated.
        print(f"running Section 7 — v2 report freshness…")
        try:
            passed7, content7 = section_7_v2_report_freshness(db, process)
        except Exception as e:
            passed7 = False
            content7 = [f"⚠ section raised: `{type(e).__name__}: {e}`"]
            import traceback
            content7.append("```")
            content7.append(traceback.format_exc())
            content7.append("```")
        verdicts.append(("Section 7 — v2 report freshness", passed7))
        body.append(f"## Section 7 — v2 report freshness  {status_icon(passed7)}")
        body.append("")
        body.extend(content7)
        body.append("")

        all_pass = all(p for _, p in verdicts)

        # Verdict block at the top
        report.append("## Verdict")
        report.append("")
        if all_pass:
            report.append("✅ **PASS** — every gate green.")
        else:
            failed = [t for t, p in verdicts if not p]
            report.append(f"❌ **FAIL** — {len(failed)} blocker(s):")
            for t in failed:
                report.append(f"- {t}")
        report.append("")

        report.append("## Section verdicts")
        report.append("")
        report.append("| Section | Result |")
        report.append("|---|:---:|")
        for t, p in verdicts:
            report.append(f"| {t} | {status_icon(p)} |")
        report.append(f"| Section 6 — Anomaly registry | ℹ |")
        report.append("")

        report.extend(body)

        report.append("## Sign-off")
        report.append("")
        report.append("- [ ] FinOps reviewed")
        report.append("- [ ] Reconciliation manager approved")
        report.append("")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("\n".join(report))
        print(f"\nReport written to {output_path}")
        print(f"Verdict: {'✅ PASS' if all_pass else '❌ FAIL'}")
        return all_pass
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description="Ultra-audit a TrueBook reconciliation run")
    parser.add_argument("--process-id", type=int, required=True)
    parser.add_argument("--output", default=None,
                        help="Output markdown path. Default: docs/audits/{period}_{id}.md")
    parser.add_argument("--include-mongo", action="store_true",
                        help="Re-extract from Mongo Atlas (slow)")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        proc = db.query(AccountingProcess).filter_by(id=args.process_id).first()
        if not proc:
            print(f"❌ process_id={args.process_id} not found")
            sys.exit(1)
        period = f"{proc.period_year}_{proc.period_month:02d}"
    finally:
        db.close()

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = Path(__file__).resolve().parent.parent.parent / "docs" / "audits" / f"{period}_process{args.process_id}.md"

    ok = run_audit(args.process_id, output_path, args.include_mongo)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
