import io
from collections import defaultdict
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONTHS_ES_UPPER = [
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
]


def _num(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _safe_str(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    return str(value)


def _parse_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None


def _date_label(value: Any) -> str:
    d = _parse_date(value)
    if d:
        return d.isoformat()
    return _safe_str(value, "")


def _month_name_upper(period_month: int) -> str:
    idx = max(1, min(12, int(period_month))) - 1
    return MONTHS_ES_UPPER[idx]


def _styled_header_row(ws, row_idx: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, max_width: int = 52):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)


def _save_workbook(wb: Workbook) -> bytes:
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _extract_fees_components(fees_result) -> Dict[str, Any]:
    merchant_summary = fees_result.merchant_summary or []
    daily_breakdown = fees_result.daily_breakdown or []
    withdrawals_summary = fees_result.withdrawals_summary or []
    refunds_summary = fees_result.refunds_summary or []
    other_fees_summary = fees_result.other_fees_summary or []
    return {
        "merchant_summary": merchant_summary,
        "daily_breakdown": daily_breakdown,
        "withdrawals_summary": withdrawals_summary,
        "refunds_summary": refunds_summary,
        "other_fees_summary": other_fees_summary,
    }


def build_fees_export(process, fees_result) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    components = _extract_fees_components(fees_result)
    merchant_summary = components["merchant_summary"]
    daily_breakdown = components["daily_breakdown"]
    withdrawals_summary = components["withdrawals_summary"]
    refunds_summary = components["refunds_summary"]
    other_fees_summary = components["other_fees_summary"]

    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "Detalle por Merchant"
    ws_summary = wb.create_sheet("Resumen por Merchant")
    ws_razon = wb.create_sheet("Resumen por Razon Social")
    ws_daily = wb.create_sheet("Tonder Fees desglose diario")

    # ---- Detalle por Merchant ----
    ws_detail["A2"] = f"REPORTE DE FEES — {period_text}   |   DETALLE POR MERCHANT"
    ws_detail["A3"] = (
        f"Montos en MXN  |  UTC-6  |  01-{_month_name_upper(process.period_month)[:3].title()}-{process.period_year} "
        f"00:00 → Fin de mes 23:59  |  Neto = Monto Procesado - Fee c/IVA"
    )
    detail_headers = [
        "",
        "Merchant",
        "Concepto",
        "Adquirente",
        "# Eventos",
        "Monto Procesado",
        "Fee %",
        "Fee Fijo",
        "Total Fee s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_detail.append([])
    ws_detail.append(detail_headers)
    _styled_header_row(ws_detail, ws_detail.max_row, 1, len(detail_headers))

    grouped_tx = defaultdict(lambda: {"events": 0, "amount": 0.0, "fee": 0.0})
    merchant_name_map: Dict[str, str] = {}
    for row in daily_breakdown:
        merchant_id = _safe_str(row.get("merchant_id"), "unknown")
        merchant_name = _safe_str(row.get("merchant_name"), merchant_id)
        acquirer = _safe_str(row.get("acquirer"), "Operativa")
        key = (merchant_id, merchant_name, f"{acquirer} - Operativa", acquirer)
        grouped_tx[key]["events"] += 1
        grouped_tx[key]["amount"] += _num(row.get("amount"))
        grouped_tx[key]["fee"] += _num(row.get("fee_amount"))
        merchant_name_map[merchant_id] = merchant_name

    w_map = { _safe_str(w.get("merchant_id"), "unknown"): w for w in withdrawals_summary }
    r_map = { _safe_str(r.get("merchant_id"), "unknown"): r for r in refunds_summary }

    all_merchants = set(merchant_name_map.keys()) | set(w_map.keys()) | set(r_map.keys())
    all_merchants |= { _safe_str(m.get("merchant_id"), "unknown") for m in merchant_summary }

    for mid in sorted(all_merchants, key=lambda x: merchant_name_map.get(x, x)):
        merchant_name = merchant_name_map.get(mid, mid)
        ws_detail.append([None, merchant_name])

        for (g_mid, _m_name, concept, acquirer), val in sorted(grouped_tx.items(), key=lambda x: (x[0][1], x[0][2])):
            if g_mid != mid:
                continue
            amount = round(val["amount"], 6)
            total_fee = round(val["fee"], 6)
            fee_pct = round((total_fee / amount) if amount else 0.0, 6)
            iva = round(total_fee * 0.16, 6)
            total_c_iva = round(total_fee + iva, 6)
            neto = round(amount - total_c_iva, 6)
            ws_detail.append([
                None,
                None,
                concept,
                acquirer,
                val["events"],
                amount,
                fee_pct,
                0.0,
                total_fee,
                iva,
                total_c_iva,
                neto,
            ])

        if mid in w_map:
            w = w_map[mid]
            total_fee = round(_num(w.get("total_fee")), 6)
            amount = round(_num(w.get("total_amount")), 6)
            iva = round(total_fee * 0.16, 6)
            ws_detail.append([
                None,
                None,
                "Withdrawals",
                "N/A",
                int(_num(w.get("count"))),
                amount if amount else "—",
                round((total_fee / amount) if amount else 0.0, 6),
                0.0,
                total_fee,
                iva,
                round(total_fee + iva, 6),
                "—",
            ])

        if mid in r_map:
            r = r_map[mid]
            total_fee = round(_num(r.get("total_fee")), 6)
            amount = round(_num(r.get("total_amount")), 6)
            iva = round(total_fee * 0.16, 6)
            ws_detail.append([
                None,
                None,
                "Autorefunds/Refunds",
                "N/A",
                int(_num(r.get("count"))),
                amount if amount else "—",
                round((total_fee / amount) if amount else 0.0, 6),
                0.0,
                total_fee,
                iva,
                round(total_fee + iva, 6),
                "—",
            ])

    # ---- Resumen por Merchant ----
    ws_summary["A2"] = f"RESUMEN DE FEES POR MERCHANT — {period_text}"
    ws_summary["A3"] = "Montos en MXN  |  Neto = Monto Procesado - Fee c/IVA"
    ws_summary.append([])
    summary_headers = [
        "",
        "Merchant",
        "Monto Procesado",
        "Fees Transacc.",
        "Other Fees",
        "Settlement",
        "Withdrawals",
        "Autorefunds",
        "Routing Fee",
        "Total s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_summary.append(summary_headers)
    _styled_header_row(ws_summary, ws_summary.max_row, 1, len(summary_headers))

    other_by_mid = defaultdict(float)
    settlement_by_mid = defaultdict(float)
    routing_by_mid = defaultdict(float)
    for item in other_fees_summary if isinstance(other_fees_summary, list) else []:
        mid = _safe_str(item.get("merchant_id"), "unknown")
        amount = _num(item.get("total_fee", item.get("amount", 0)))
        concept = _safe_str(item.get("concept", item.get("type", "other"))).lower()
        if "settlement" in concept:
            settlement_by_mid[mid] += amount
        elif "routing" in concept:
            routing_by_mid[mid] += amount
        else:
            other_by_mid[mid] += amount

    merchant_ids = [ _safe_str(m.get("merchant_id"), "unknown") for m in merchant_summary ]
    merchant_rows = { _safe_str(m.get("merchant_id"), "unknown"): m for m in merchant_summary }
    for mid in sorted(set(merchant_ids) | set(w_map.keys()) | set(r_map.keys())):
        m = merchant_rows.get(mid, {})
        name = _safe_str(m.get("merchant_name"), mid)
        gross = round(_num(m.get("gross_amount")), 6)
        tx_fee = round(_num(m.get("total_fee")), 6)
        withdrawals = round(_num((w_map.get(mid) or {}).get("total_fee")), 6)
        autorefunds = round(_num((r_map.get(mid) or {}).get("total_fee")), 6)
        other_fees = round(other_by_mid[mid], 6)
        settlement = round(settlement_by_mid[mid], 6)
        routing = round(routing_by_mid[mid], 6)
        total_s_iva = round(tx_fee + other_fees + settlement + withdrawals + autorefunds + routing, 6)
        iva = round(total_s_iva * 0.16, 6)
        total_c_iva = round(total_s_iva + iva, 6)
        neto = round(gross - total_c_iva, 6) if gross else "—"
        ws_summary.append([
            None,
            name,
            gross,
            tx_fee,
            other_fees,
            settlement,
            withdrawals,
            autorefunds,
            routing if routing else "—",
            total_s_iva,
            iva,
            total_c_iva,
            neto,
        ])

    # ---- Resumen por Razon Social ----
    ws_razon["A2"] = f"RESUMEN DE FEES POR RAZON SOCIAL — {period_text}"
    ws_razon["A3"] = "Montos en MXN  |  Consolidado por razón social (placeholder por merchant)"
    ws_razon.append([])
    razon_headers = [
        "",
        "Razon Social",
        "Merchants",
        "Monto Procesado",
        "Fees Transacc.",
        "Other Fees",
        "Settlement",
        "Withdrawals",
        "Routing Fee",
        "Total s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_razon.append(razon_headers)
    _styled_header_row(ws_razon, ws_razon.max_row, 1, len(razon_headers))

    for row in ws_summary.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        ws_razon.append([
            None,
            row[1],  # Razon Social (sin catálogo aún)
            row[1],  # Merchants
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[8],
            row[9],
            row[10],
            row[11],
            row[12],
        ])

    # ---- Desglose diario ----
    ws_daily["A1"] = (
        f"TONDER — FEES {period_text} | DESGLOSE DIARIO | Montos en MXN | UTC-6 | "
        f"01-{_month_name_upper(process.period_month)[:3].title()}-{process.period_year} → Fin de mes"
    )
    daily_headers = [
        "Fecha",
        "Merchant",
        "Concepto / Operativa",
        "# Eventos",
        "Monto Procesado",
        "Fee s/IVA",
        "IVA (16%)",
        "Total c/IVA",
    ]
    ws_daily.append(daily_headers)
    _styled_header_row(ws_daily, 2, 1, len(daily_headers))

    daily_grouped = defaultdict(lambda: {"events": 0, "amount": 0.0, "fee": 0.0})
    for row in daily_breakdown:
        d = _date_label(row.get("date"))
        merchant = _safe_str(row.get("merchant_name"), _safe_str(row.get("merchant_id"), "unknown"))
        acquirer = _safe_str(row.get("acquirer"), "Operativa")
        concept = f"{acquirer}-Operativa"
        key = (d, merchant, concept)
        daily_grouped[key]["events"] += 1
        daily_grouped[key]["amount"] += _num(row.get("amount"))
        daily_grouped[key]["fee"] += _num(row.get("fee_amount"))

    last_date = None
    for (d, merchant, concept), val in sorted(daily_grouped.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        if d != last_date:
            pretty = d
            parsed = _parse_date(d)
            if parsed:
                pretty = parsed.strftime("  %d/%m/%Y")
            ws_daily.append([pretty])
            last_date = d
        fee = round(val["fee"], 6)
        iva = round(fee * 0.16, 6)
        ws_daily.append([
            None,
            merchant,
            concept,
            val["events"],
            round(val["amount"], 6),
            fee,
            iva,
            round(fee + iva, 6),
        ])

    for ws in [ws_detail, ws_summary, ws_razon, ws_daily]:
        _autowidth(ws)

    filename = f"FEES_{_month_name_upper(process.period_month)}_{process.period_year}_FINAL.xlsx"
    return filename, _save_workbook(wb)


def build_kushki_export(process, kushki_result) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    daily_summary = kushki_result.daily_summary or []
    merchant_detail = kushki_result.merchant_detail or []

    wb = Workbook()
    ws_daily = wb.active
    ws_daily.title = "Resumen Diario"
    ws_detail = wb.create_sheet("Detalle por Merchant")
    ws_pivot = wb.create_sheet("Pivot por Merchant")

    ws_daily["A1"] = f"KUSHKI — RESUMEN DIARIO DE LIQUIDACIONES · {period_text}"
    daily_headers = [
        "Fecha Liq.",
        "# Txns",
        "Monto Bruto",
        "Comisión Kushki",
        "IVA Kushki",
        "RR Retenido",
        "RR Liberado",
        "Com. Tonder s/IVA",
        "Ajustes",
        "Depósito Neto (Abonar)",
    ]
    ws_daily.append(daily_headers)
    _styled_header_row(ws_daily, 2, 1, len(daily_headers))

    for row in sorted(daily_summary, key=lambda r: _date_label(r.get("date"))):
        gross = round(_num(row.get("gross_amount")), 6)
        commission = round(_num(row.get("commission")), 6)
        iva = round(commission * 0.16, 6)
        rolling = round(_num(row.get("rolling_reserve")), 6)
        net = round(_num(row.get("net_deposit")), 6)
        ws_daily.append([
            _date_label(row.get("date")),
            int(_num(row.get("tx_count"))),
            gross,
            commission,
            iva,
            rolling,
            0.0,
            0.0,
            0.0,
            net,
        ])

    ws_detail["A1"] = f"KUSHKI — DETALLE POR MERCHANT · {period_text}"
    detail_headers = [
        "Fecha Liq.",
        "Merchant",
        "# Txns",
        "Monto Bruto",
        "Com. Kushki + IVA",
        "Rolling Reserve",
        "RR Liberado",
        "Com. Tonder %",
        "Depósito Neto Merchant",
    ]
    ws_detail.append(detail_headers)
    _styled_header_row(ws_detail, 2, 1, len(detail_headers))

    for row in sorted(merchant_detail, key=lambda r: _safe_str(r.get("merchant_name"), "")):
        ws_detail.append([
            f"{process.period_year}-{str(process.period_month).zfill(2)}",
            _safe_str(row.get("merchant_name"), "unknown"),
            int(_num(row.get("tx_count"))),
            round(_num(row.get("gross_amount")), 6),
            round(_num(row.get("commission")), 6),
            round(_num(row.get("rolling_reserve")), 6),
            0.0,
            0.0,
            round(_num(row.get("net_deposit")), 6),
        ])

    ws_pivot["A1"] = f"KUSHKI — ACUMULADO POR MERCHANT · {period_text}"
    pivot_headers = ["Merchant", "# Txns", "Monto Bruto", "Com. Kushki + IVA", "Tasa Efectiva", "Depósito Neto"]
    ws_pivot.append(pivot_headers)
    _styled_header_row(ws_pivot, 2, 1, len(pivot_headers))

    for row in sorted(merchant_detail, key=lambda r: _safe_str(r.get("merchant_name"), "")):
        gross = round(_num(row.get("gross_amount")), 6)
        commission = round(_num(row.get("commission")), 6)
        tasa = round((commission / gross) if gross else 0.0, 12)
        ws_pivot.append([
            _safe_str(row.get("merchant_name"), "unknown"),
            int(_num(row.get("tx_count"))),
            gross,
            commission,
            tasa,
            round(_num(row.get("net_deposit")), 6),
        ])

    for ws in [ws_daily, ws_detail, ws_pivot]:
        _autowidth(ws)

    filename = f"KUSHKI_{_month_name_upper(process.period_month)}_{process.period_year}_v3.xlsx"
    return filename, _save_workbook(wb)


def _infer_category(description: str, debit: float, credit: float) -> str:
    d = (description or "").lower()
    if "kushki" in d:
        return "Kushki – Liquidación"
    if "settlement" in d or ("stp" in d and debit > 0):
        return "Settlement merchant"
    if credit > 0:
        return "Abonos"
    if debit > 0:
        return "Cargos"
    return "Otros"


def build_banregio_export(process, banregio_result, kushki_result, conciliation_results: List[Any]) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    movements = banregio_result.movements or []
    summary = banregio_result.summary or {}
    daily_summary = (kushki_result.daily_summary if kushki_result else []) or []

    kvb = None
    for c in conciliation_results or []:
        if getattr(c, "conciliation_type", "") == "kushki_vs_banregio":
            kvb = c
            break

    matched = (kvb.matched if kvb else []) or []
    unmatched_kushki = (kvb.unmatched_kushki if kvb else []) or []
    unmatched_banregio = (kvb.unmatched_banregio if kvb else []) or []
    total_difference = _num(getattr(kvb, "total_difference", 0)) if kvb else 0.0

    wb = Workbook()
    ws_moves = wb.active
    ws_moves.title = "MOVIMIENTOS"
    ws_summary = wb.create_sheet("RESUMEN")
    ws_cross = wb.create_sheet("CRUCE KUSHKI")

    # ---- MOVIMIENTOS ----
    ws_moves["A1"] = f"ESTADO DE CUENTA BANREGIO – {period_text}   |   TRES COMAS S.A.P.I. DE C.V.   |   Cta. 001-9"
    ws_moves["A2"] = (
        f"Total Abonos: ${_num(summary.get('total_credits')):,.2f}  |  "
        f"Total Cargos: ${_num(summary.get('total_debits')):,.2f}  |  "
        f"Neto: ${_num(summary.get('net')):,.2f}"
    )
    move_headers = ["DÍA", "FECHA", "TIPO", "CONCEPTO / CONTRAPARTE", "CATEGORÍA", "CARGOS (MXN)", "ABONOS (MXN)", "SALDO (MXN)"]
    ws_moves.append(move_headers)
    _styled_header_row(ws_moves, 3, 1, len(move_headers))

    category_acc = defaultdict(lambda: {"debits": 0.0, "credits": 0.0, "count": 0})
    running_balance = 0.0
    for mv in movements:
        d = _parse_date(mv.get("date"))
        debit = round(_num(mv.get("debit")), 6)
        credit = round(_num(mv.get("credit")), 6)
        desc = _safe_str(mv.get("description"), "")
        typ = _safe_str(mv.get("type"), "INT" if ("spei" in desc.lower() or "traspaso" in desc.lower()) else "")
        category = _safe_str(mv.get("category"), _infer_category(desc, debit, credit))
        running_balance += (credit - debit)
        category_acc[category]["debits"] += debit
        category_acc[category]["credits"] += credit
        category_acc[category]["count"] += 1
        ws_moves.append([
            d.day if d else None,
            d if d else _safe_str(mv.get("date"), ""),
            typ,
            desc,
            category,
            debit if debit > 0 else None,
            credit if credit > 0 else None,
            round(running_balance, 6),
        ])

    # ---- RESUMEN ----
    ws_summary["A1"] = f"RESUMEN {period_text} – BANREGIO TRES COMAS"
    summary_headers = ["CATEGORÍA", "TOTAL CARGOS", "TOTAL ABONOS", "# MOVS"]
    ws_summary.append(summary_headers)
    _styled_header_row(ws_summary, 2, 1, len(summary_headers))
    for category, data in sorted(category_acc.items(), key=lambda x: x[0]):
        ws_summary.append([
            category,
            round(data["debits"], 6) if data["debits"] else None,
            round(data["credits"], 6) if data["credits"] else None,
            data["count"],
        ])

    # ---- CRUCE KUSHKI ----
    ws_cross["A1"] = f"CRUCE KUSHKI vs BANREGIO – LIQUIDACIONES {period_text}   |   TRES COMAS S.A.P.I. DE C.V."
    total_expected = len([d for d in daily_summary if _num(d.get("net_deposit")) > 0])
    total_matched = len(matched)
    total_diff_rows = len(unmatched_kushki) + len(unmatched_banregio)
    ws_cross["A2"] = (
        f"{total_expected} depósitos esperados · {total_matched} conciliados · "
        f"{total_diff_rows} diferencias · Diferencia total: ${total_difference:,.2f}"
    )
    cross_headers = [
        "FECHA LIQ.",
        "# TXNS",
        "MONTO BRUTO",
        "COM. KUSHKI + IVA",
        "RR RETENIDO",
        "RR LIBERADO",
        "DEP. NETO KUSHKI",
        "ABONO BANREGIO",
        "DIFERENCIA",
        "ESTADO",
        None,
        "RESUMEN DEL CRUCE",
        None,
    ]
    ws_cross.append(cross_headers)
    _styled_header_row(ws_cross, 3, 1, 10)
    _styled_header_row(ws_cross, 3, 12, 13)

    matched_by_date = { _date_label(m.get("date")): m for m in matched }
    row_ptr = 4
    for daily in sorted(daily_summary, key=lambda r: _date_label(r.get("date"))):
        kushki_amount = round(_num(daily.get("net_deposit")), 6)
        if kushki_amount <= 0:
            continue
        date_key = _date_label(daily.get("date"))
        m = matched_by_date.get(date_key)
        banregio_amount = round(_num(m.get("banregio_amount")), 6) if m else 0.0
        diff = round(abs(kushki_amount - banregio_amount), 6) if m else kushki_amount
        ws_cross.append([
            _parse_date(date_key) or date_key,
            int(_num(daily.get("tx_count"))),
            round(_num(daily.get("gross_amount")), 6),
            round(_num(daily.get("commission")), 6),
            round(_num(daily.get("rolling_reserve")), 6),
            0.0,
            kushki_amount,
            banregio_amount if m else None,
            diff if diff else None,
            "✓ OK" if m and diff <= 0.01 else "⚠ Revisar",
            None,
            None,
            None,
        ])
        row_ptr += 1

    cross_total_kushki = round(sum(_num(d.get("net_deposit")) for d in daily_summary), 6)
    cross_total_banregio = round(sum(_num(m.get("banregio_amount")) for m in matched), 6)
    side_rows = [
        ("Depósitos Kushki", cross_total_kushki),
        ("Abonos Banregio", cross_total_banregio),
        ("Diferencia total", round(abs(cross_total_kushki - cross_total_banregio), 6)),
        ("Días conciliados", total_matched),
        ("Días con diferencia", total_diff_rows),
    ]
    side_start = 4
    for i, (label, value) in enumerate(side_rows):
        ws_cross.cell(row=side_start + i, column=12, value=label)
        ws_cross.cell(row=side_start + i, column=13, value=value)

    for ws in [ws_moves, ws_summary, ws_cross]:
        _autowidth(ws)

    filename = f"BANREGIO_{_month_name_upper(process.period_month)}_{process.period_year}_CONCILIADO_v2.xlsx"
    return filename, _save_workbook(wb)


# ═══════════════════════════════════════════════════════════════════════
# RECONCILIATION EXPORT — Full reconciliation view with checkmarks
# ═══════════════════════════════════════════════════════════════════════

CLS_LABELS = {
    "kushki_acquirer": "Kushki",
    "bitso_acquirer": "Bitso",
    "unlimit_acquirer": "Unlimit",
    "pagsmile_acquirer": "Pagsmile",
    "stp_acquirer": "STP",
    "settlement_to_merchant": "Dispersión a comercio",
    "revenue": "Revenue Tonder",
    "investment": "Inversión",
    "tax": "ISR",
    "bank_expense": "Comisión bancaria",
    "currency_sale": "Venta de divisas",
    "transfer_between_accounts": "Traspaso entre cuentas",
    "unclassified": "SIN CLASIFICAR",
    "ignored": "Ignorado",
}


def build_reconciliation_export(
    process,
    movements_data: list,
    summary: dict,
    acquirer_data: dict,
    alerts: list,
) -> tuple:
    """
    Export the full reconciliation view as Excel.
    Beautiful, branded design with TrueBook colors.

    Sheets:
    1. Reconciliación — every movement with checkmark + classification
    2. Por Adquirente — breakdown per acquirer with merchants
    3. Alertas — pending items and reconciliation alerts
    """
    wb = Workbook()
    month_name = _month_name_upper(process.period_month)
    year = process.period_year
    bank = getattr(process, "bank_account", "Banregio") or "Banregio"

    # ── Palette: clean grey + blue ──
    DARK = "1F2937"         # gray-800
    MID = "374151"          # gray-700
    BLUE = "2563EB"         # blue-600
    BLUE_DARK = "1E40AF"    # blue-800
    BLUE_LIGHT = "EFF6FF"   # blue-50
    BLUE_MED = "DBEAFE"     # blue-100
    WHITE = "FFFFFF"
    GRAY_50 = "F9FAFB"
    GRAY_100 = "F3F4F6"
    GRAY_200 = "E5E7EB"
    GRAY_300 = "D1D5DB"
    GRAY_500 = "6B7280"
    GRAY_900 = "111827"
    RED_LIGHT = "FEF2F2"
    RED_TX = "B91C1C"
    GREEN_TX = "047857"
    AMBER_TX = "92400E"
    AMBER_LIGHT = "FFFBEB"

    # ── Reusable styles ──
    title_font = Font(name="Calibri", bold=True, size=13, color=DARK)
    subtitle_font = Font(name="Calibri", size=10, color=GRAY_500)
    header_font = Font(name="Calibri", bold=True, size=9, color=GRAY_500)
    header_fill = PatternFill("solid", fgColor=GRAY_50)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Calibri", size=10, color=MID)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    green_fill = PatternFill("solid", fgColor=WHITE)
    red_fill = PatternFill("solid", fgColor=RED_LIGHT)
    amber_fill = PatternFill("solid", fgColor=AMBER_LIGHT)
    gray_fill = PatternFill("solid", fgColor=GRAY_100)
    light_fill = PatternFill("solid", fgColor=GRAY_50)
    white_fill = PatternFill("solid", fgColor=WHITE)
    from openpyxl.styles import Border, Side
    thin_border = Border(
        bottom=Side(style="thin", color=GRAY_200),
    )
    header_border = Border(
        bottom=Side(style="medium", color=GRAY_300),
    )

    def _brand_header(ws, row_idx, start, end):
        for col in range(start, end + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

    def _brand_title(ws, row_idx, text, merge_to=6):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=merge_to)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = text
        cell.font = title_font
        cell.alignment = Alignment(vertical="center")

    def _brand_subtitle(ws, row_idx, text, merge_to=6):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=merge_to)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = text
        cell.font = subtitle_font

    def _section_header(ws, row_idx, text, cols=5):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=cols)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = text
        cell.font = Font(name="Calibri", bold=True, size=10, color=BLUE_DARK)
        cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        cell.alignment = Alignment(vertical="center")
        for col in range(2, cols + 1):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT)

    cov = summary.get("coverage_pct", 0)
    classified = summary.get("classified", 0)
    total = summary.get("total_movements", 0)
    unclassified = summary.get("unclassified", 0)

    # ── Sheet 1: Reconciliación ───────────────────────────────────────
    ws = wb.active
    ws.title = "Reconciliación"
    ws.sheet_properties.tabColor = BLUE

    # Title block
    _brand_title(ws, 1, f"TrueBook — Reconciliación {bank}", 10)
    _brand_subtitle(ws, 2, f"{month_name} {year}  |  {process.name}", 10)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # KPI row
    kpi_row = 4
    kpi_data = [
        ("Movimientos", total, DARK),
        ("Reconciliados", classified, GREEN_TX),
        ("Pendientes", unclassified, RED_TX if unclassified > 0 else GRAY_500),
        ("Cobertura", f"{cov}%", BLUE),
    ]
    for i, (label, val, color) in enumerate(kpi_data):
        col = 1 + i * 2
        ws.cell(row=kpi_row, column=col, value=label).font = Font(name="Calibri", size=9, color=GRAY_500)
        c = ws.cell(row=kpi_row + 1, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=14, color=color)
    ws.row_dimensions[kpi_row + 1].height = 24

    # Headers
    h_row = 7
    headers = ["", "Fecha", "Descripción", "Cargo", "Abono", "Saldo", "Clasificación", "Adquirente", "Método", "Estado"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=h_row, column=i, value=h)
    _brand_header(ws, h_row, 1, len(headers))
    ws.row_dimensions[h_row].height = 22
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['C'].width = 45

    # Compute opening balance from summary if available, otherwise 0
    opening_balance = _num(summary.get("opening_balance", 0))

    # Opening Balance row
    ob_row = h_row + 1
    ws.cell(row=ob_row, column=3, value="SALDO INICIAL (Opening Balance)").font = Font(
        name="Calibri", bold=True, size=10, color=BLUE_DARK)
    ws.cell(row=ob_row, column=6, value=opening_balance)
    ws.cell(row=ob_row, column=6).number_format = money_fmt
    ws.cell(row=ob_row, column=6).font = Font(name="Calibri", bold=True, size=10, color=BLUE_DARK)
    for col in range(1, len(headers) + 1):
        ws.cell(row=ob_row, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        ws.cell(row=ob_row, column=col).border = thin_border

    # Data rows
    running_balance = opening_balance
    for r_idx, m in enumerate(movements_data, h_row + 2):
        is_recon = m.get("is_reconciled", False)
        cls_label = CLS_LABELS.get(m.get("classification", ""), m.get("classification", ""))

        debit = _num(m.get("debit", 0))
        credit = _num(m.get("credit", 0))
        running_balance += (credit - debit)

        ws.cell(row=r_idx, column=1, value="✓" if is_recon else "✗").font = Font(
            name="Calibri", bold=True, size=10, color=BLUE if is_recon else RED_TX)
        ws.cell(row=r_idx, column=2, value=m.get("date", "")).font = body_font
        ws.cell(row=r_idx, column=3, value=m.get("description", "")).font = body_font
        ws.cell(row=r_idx, column=4, value=debit if debit > 0 else None)
        ws.cell(row=r_idx, column=5, value=credit if credit > 0 else None)
        ws.cell(row=r_idx, column=6, value=round(running_balance, 2))
        ws.cell(row=r_idx, column=6).number_format = money_fmt
        ws.cell(row=r_idx, column=6).font = Font(name="Calibri", size=10, color=DARK)
        ws.cell(row=r_idx, column=7, value=cls_label).font = Font(name="Calibri", size=9, color=BLUE_DARK)
        ws.cell(row=r_idx, column=8, value=m.get("acquirer") or "").font = Font(name="Calibri", size=9, color=GRAY_500)
        ws.cell(row=r_idx, column=9, value=m.get("method") or "").font = Font(name="Calibri", size=9, color=GRAY_300)
        estado_cell = ws.cell(row=r_idx, column=10, value="Reconciliado" if is_recon else "PENDIENTE")
        estado_cell.font = Font(name="Calibri", bold=True, size=9,
                                color=BLUE if is_recon else RED_TX)

        # Clean alternating rows — pending gets subtle red, reconciled alternates white/gray
        row_fill = white_fill
        if not is_recon:
            row_fill = red_fill
        elif r_idx % 2 == 0:
            row_fill = light_fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in (4, 5):
                cell.number_format = money_fmt
                if col == 4 and cell.value:
                    cell.font = Font(name="Calibri", size=10, color=RED_TX)
                elif col == 5 and cell.value:
                    cell.font = Font(name="Calibri", size=10, color=MID)

    # Closing Balance row
    cb_row = h_row + 2 + len(movements_data)
    ws.cell(row=cb_row, column=3, value="SALDO FINAL (Closing Balance)").font = Font(
        name="Calibri", bold=True, size=10, color=BLUE_DARK)
    ws.cell(row=cb_row, column=6, value=round(running_balance, 2))
    ws.cell(row=cb_row, column=6).number_format = money_fmt
    ws.cell(row=cb_row, column=6).font = Font(name="Calibri", bold=True, size=10, color=BLUE_DARK)
    for col in range(1, len(headers) + 1):
        ws.cell(row=cb_row, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        ws.cell(row=cb_row, column=col).border = thin_border

    _autowidth(ws, max_width=50)

    # ── Sheet 2: Por Adquirente ───────────────────────────────────────
    ws2 = wb.create_sheet("Por Adquirente")
    ws2.sheet_properties.tabColor = BLUE_DARK

    _brand_title(ws2, 1, f"TrueBook — Desglose por Adquirente", 17)
    _brand_subtitle(ws2, 2, f"{month_name} {year}  |  {bank}  |  {process.name}", 17)

    row = 4
    acquirers = acquirer_data.get("acquirers", [])
    for acq in acquirers:
        name = acq.get("name", "").upper()
        deps = acq.get("deposits", [])
        total_amt = acq.get("total_amount", 0)

        # Acquirer section header — blue-light bar
        _section_header(ws2, row, f"  {name}  —  {len(deps)} depósitos  —  ${total_amt:,.2f} MXN", 5)
        ws2.row_dimensions[row].height = 26
        row += 1

        # Deposit headers
        dep_headers = ["Fecha", "Descripción", "Monto"]
        for i, h in enumerate(dep_headers, 1):
            ws2.cell(row=row, column=i, value=h)
        _brand_header(ws2, row, 1, len(dep_headers))
        row += 1

        for dep in deps:
            ws2.cell(row=row, column=1, value=dep.get("date", "")).font = body_font
            ws2.cell(row=row, column=2, value=dep.get("description", "")).font = body_font
            c = ws2.cell(row=row, column=3, value=dep.get("amount", 0))
            c.number_format = money_fmt
            c.font = Font(name="Calibri", size=10, color=MID)
            for col in range(1, 4):
                ws2.cell(row=row, column=col).border = thin_border
            row += 1

        # Merchant detail if available
        merchants = acq.get("merchants", [])
        if merchants:
            row += 1
            ws2.cell(row=row, column=1, value="Desglose por comercio").font = Font(
                name="Calibri", bold=True, size=9, color=GRAY_500)
            row += 1
            merch_headers = [
                "Comercio", "# Txns", "Monto Bruto", "Bruto Ajustes",
                "Com. Kushki", "IVA Kushki", "Com. Kushki + IVA",
                "RR Retenido", "Devolución", "Contracargo", "Cancelación",
                "Manual", "RR Liberado", "Depósito Neto",
                "Com. Tonder s/IVA", "IVA (16%)", "Com. Tonder c/IVA",
            ]
            for i, h in enumerate(merch_headers, 1):
                ws2.cell(row=row, column=i, value=h)
            _brand_header(ws2, row, 1, len(merch_headers))
            ws2.row_dimensions[row].height = 22
            row += 1
            merch_fields = [
                "merchant_name", "tx_count", "gross_amount", "adjustments",
                "kushki_commission", "iva_kushki_commission", "commission",
                "rolling_reserve", "refund", "chargeback", "void",
                "manual_adj", "rr_released", "net_deposit",
                "tonder_fee", "tonder_iva", "tonder_fee_iva",
            ]
            for m_idx, m in enumerate(sorted(merchants, key=lambda x: _num(x.get("net_deposit", 0)), reverse=True)):
                for ci, field in enumerate(merch_fields, 1):
                    val = m.get(field, 0)
                    cell = ws2.cell(row=row, column=ci)
                    if field == "merchant_name":
                        cell.value = val or ""
                        cell.font = Font(name="Calibri", bold=True, size=10, color=DARK)
                    elif field == "tx_count":
                        cell.value = int(_num(val))
                        cell.number_format = '#,##0'
                        cell.font = body_font
                    else:
                        cell.value = _num(val)
                        cell.number_format = money_fmt
                        cell.font = body_font
                        if field == "net_deposit":
                            cell.font = Font(name="Calibri", bold=True, size=10, color=BLUE_DARK)
                        elif field == "chargeback" and _num(val) < 0:
                            cell.font = Font(name="Calibri", size=10, color=RED_TX)
                        elif field == "commission":
                            cell.font = Font(name="Calibri", size=10, color=GRAY_500)
                    cell.border = thin_border
                    if m_idx % 2 == 1:
                        cell.fill = light_fill
                row += 1

        row += 2  # spacing between acquirers

    _autowidth(ws2, max_width=18)

    # ── Sheet 3: Alertas ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Alertas")
    ws3.sheet_properties.tabColor = GRAY_500

    _brand_title(ws3, 1, "TrueBook — Alertas de Reconciliación")
    _brand_subtitle(ws3, 2, f"{month_name} {year}  |  Cobertura: {cov}%  |  {unclassified} pendientes")

    # Pending movements
    ws3.cell(row=4, column=1, value="MOVIMIENTOS PENDIENTES DE RECONCILIAR")
    ws3.cell(row=4, column=1).font = Font(name="Calibri", bold=True, size=11, color=DARK)

    pend_headers = ["#", "Fecha", "Descripción", "Cargo", "Abono", "Acción requerida"]
    for i, h in enumerate(pend_headers, 1):
        ws3.cell(row=5, column=i, value=h)
    _brand_header(ws3, 5, 1, len(pend_headers))

    pend_row = 6
    pend_count = 0
    for m in movements_data:
        if not m.get("is_reconciled", False):
            pend_count += 1
            ws3.cell(row=pend_row, column=1, value=pend_count)
            ws3.cell(row=pend_row, column=2, value=m.get("date", ""))
            ws3.cell(row=pend_row, column=3, value=m.get("description", ""))
            ws3.cell(row=pend_row, column=4, value=m.get("debit") if m.get("debit", 0) > 0 else None)
            ws3.cell(row=pend_row, column=5, value=m.get("credit") if m.get("credit", 0) > 0 else None)
            ws3.cell(row=pend_row, column=6, value="Clasificar manualmente o ejecutar Warren AI")
            for col in [4, 5]:
                ws3.cell(row=pend_row, column=col).number_format = '#,##0.00'
            for col in range(1, len(pend_headers) + 1):
                ws3.cell(row=pend_row, column=col).fill = red_fill
            pend_row += 1

    # System alerts
    if alerts:
        pend_row += 2
        ws3.cell(row=pend_row, column=1, value="ALERTAS DEL SISTEMA")
        ws3.cell(row=pend_row, column=1).font = Font(name="Calibri", bold=True, size=11, color=DARK)
        pend_row += 1
        alert_headers = ["Nivel", "Tipo", "Título", "Mensaje"]
        for i, h in enumerate(alert_headers, 1):
            ws3.cell(row=pend_row, column=i, value=h)
        _brand_header(ws3, pend_row, 1, len(alert_headers))
        pend_row += 1
        for a in alerts:
            ws3.cell(row=pend_row, column=1, value=a.get("alert_level", "")).font = Font(
                name="Calibri", bold=True, size=9, color=GRAY_500)
            ws3.cell(row=pend_row, column=2, value=a.get("alert_type", "")).font = body_font
            ws3.cell(row=pend_row, column=3, value=a.get("title", "")).font = Font(
                name="Calibri", bold=True, size=10, color=DARK)
            ws3.cell(row=pend_row, column=4, value=a.get("message", "")).font = body_font
            level = a.get("alert_level", "")
            fill = red_fill if level == "CRITICAL" else amber_fill if level in ("WARNING", "UNCLASSIFIED") else gray_fill
            for col in range(1, 5):
                ws3.cell(row=pend_row, column=col).fill = fill
                ws3.cell(row=pend_row, column=col).border = thin_border
            pend_row += 1

    _autowidth(ws3)

    filename = f"RECONCILIACION_{bank.upper()}_{month_name}_{year}.xlsx"
    return filename, _save_workbook(wb)
