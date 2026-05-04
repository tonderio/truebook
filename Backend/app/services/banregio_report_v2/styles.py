"""
Visual palette for the Banregio v2 report (spec §6).

All colors, fonts, and fills as ready-to-use openpyxl objects so the
sheet builders just import and apply — no inline hex codes scattered
across the codebase.

Reference: the example file
`/Users/yuyo/Downloads/RECONCILIACION_BANREGIO_MARZO_2026_v2.xlsx`
which encodes the spec's tokens. Inspecting it shows e.g. titles in
`#1F2937` bold size 13, subtitles `#6B7280` size 10, etc.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── color tokens (spec §6.2 + §6.3) ──────────────────────────────────────

class C:
    """Color tokens as ARGB hex strings (openpyxl wants 'FF' alpha prefix)."""
    TITLE = "FF1F2937"
    SUBTITLE = "FF6B7280"
    BODY = "FF374151"
    BLUE_BOLD = "FF1E40AF"
    BLUE = "FF2563EB"
    RED = "FFB91C1C"
    ORANGE = "FFB45309"
    GREEN = "FF166534"
    GREEN_DARK = "FF047857"
    WHITE = "FFFFFFFF"
    GRAY_LIGHT = "FFD1D5DB"


class Fill:
    """Background fill ARGB hex codes."""
    SECTION_KUSHKI = "FFEFF6FF"
    SECTION_BITSO = "FFEEF2FF"
    SECTION_OXXO = "FFECFDF5"
    SECTION_STP = "FFFDF4FF"
    SECTION_UNLIMIT = "FFFFF7ED"
    SECTION_RESUMEN = "FF1E3A5F"
    COL_HEADER = "FFF9FAFB"
    TOTAL = "FFF0F9FF"
    ALERT_RED = "FFFEF2F2"
    ALERT_YELLOW = "FFFEFCE8"
    ALERT_GREEN = "FFF0FDF4"
    ALERT_ORANGE = "FFFFF7ED"
    WHITE = "FFFFFFFF"


# ── reusable Font/Fill objects (mutable — instantiate fresh when changing) ─

FONT_NAME = "Calibri"


def font(*, size=10, color=C.BODY, bold=False, italic=False) -> Font:
    return Font(name=FONT_NAME, size=size, color=color, bold=bold, italic=italic)


def fill(hex_argb: str) -> PatternFill:
    return PatternFill(start_color=hex_argb, end_color=hex_argb, fill_type="solid")


# Pre-built common combinations
TITLE_FONT = font(size=13, color=C.TITLE, bold=True)
SUBTITLE_FONT = font(size=10, color=C.SUBTITLE)

KPI_LABEL_FONT = font(size=9, color=C.SUBTITLE)
KPI_VALUE_FONT = font(size=14, color=C.TITLE, bold=True)
KPI_VALUE_GREEN = font(size=14, color=C.GREEN_DARK, bold=True)
KPI_VALUE_BLUE = font(size=14, color=C.BLUE, bold=True)
KPI_VALUE_GRAY = font(size=14, color=C.SUBTITLE, bold=True)

COL_HEADER_FONT = font(size=9, color=C.SUBTITLE, bold=True)
COL_HEADER_FILL = fill(Fill.COL_HEADER)

BODY_FONT = font(size=10, color=C.BODY)
BODY_BOLD = font(size=10, color=C.TITLE, bold=True)
BODY_RED = font(size=10, color=C.RED, bold=True)

SECTION_FONT_BLUE = font(size=10, color=C.BLUE_BOLD, bold=True)
SECTION_FILL_BLUE = fill(Fill.SECTION_KUSHKI)

CHECK_FONT = font(size=10, color=C.BLUE, bold=True)

CLASSIFICATION_FONT = font(size=9, color=C.BLUE_BOLD)

METHOD_FONT = font(size=9, color=C.GRAY_LIGHT)

STATUS_RECONCILED_FONT = font(size=9, color=C.BLUE, bold=True)
STATUS_PENDING_FONT = font(size=9, color=C.ORANGE, bold=True)


# ── alignment ────────────────────────────────────────────────────────────

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── number formats ───────────────────────────────────────────────────────

FMT_MXN = '"$"#,##0.00'
FMT_INT = "#,##0"
FMT_PCT = "0.0%"
FMT_DATE = "DD/MM/YYYY"


# ── classification → display color ───────────────────────────────────────

# Spec §6.4 implies category-aware coloring. The example file uses
# blue_bold for Inversión; we extend the same logic to other categories
# in a way that's easy to override per FinOps preference.

CLASSIFICATION_COLOR: dict[str, str] = {
    # Acquirers — blue_bold (deposit-side accent)
    "Kushki": C.BLUE_BOLD,
    "kushki_acquirer": C.BLUE_BOLD,
    "Bitso": C.BLUE_BOLD,
    "bitso_acquirer": C.BLUE_BOLD,
    "Pagsmile": C.BLUE_BOLD,
    "pagsmile_acquirer": C.BLUE_BOLD,
    "OXXOPay": C.BLUE_BOLD,
    "OXXOPay (vía Pagsmile)": C.BLUE_BOLD,
    "STP": C.BLUE_BOLD,
    "stp_acquirer": C.BLUE_BOLD,
    "Unlimit": C.BLUE_BOLD,
    "unlimit_acquirer": C.BLUE_BOLD,

    # Operational
    "Inversión": C.BLUE_BOLD,
    "investment": C.BLUE_BOLD,
    "Dispersión a comercio": C.SUBTITLE,
    "settlement_to_merchant": C.SUBTITLE,
    "Traspaso entre cuentas": C.SUBTITLE,
    "transfer_between_accounts": C.SUBTITLE,
    "revenue": C.SUBTITLE,
    "Comisión bancaria": C.RED,
    "bank_expense": C.RED,
    "ISR": C.ORANGE,
    "tax": C.ORANGE,
    "Venta de divisas": C.SUBTITLE,
    "currency_sale": C.SUBTITLE,

    # Default
    "unclassified": C.RED,
}


# ── DB-classification → display name (spec naming) ───────────────────────

# Sheet 1 (bank-statement view) → uses the literal classification names
# from spec §2.4. Pagsmile is "Pagsmile" because that's how the bank
# describes the SPEI. Sheet 2 (acquirer cuadre) uses the longer
# OXXOPay name — see SHEET2_ACQUIRER_LABEL.
DB_TO_DISPLAY: dict[str, tuple[str, str | None]] = {
    "kushki_acquirer": ("Kushki", "kushki"),
    "bitso_acquirer": ("Bitso", "bitso"),
    "pagsmile_acquirer": ("Pagsmile", "oxxopay"),
    "stp_acquirer": ("STP", "stp"),
    "unlimit_acquirer": ("Unlimit", "unlimit"),
    "settlement_to_merchant": ("Dispersión a comercio", None),
    "investment": ("Inversión", None),
    "tax": ("ISR", None),
    "revenue": ("Traspaso entre cuentas", None),
    "transfer_between_accounts": ("Traspaso entre cuentas", None),
    "bank_expense": ("Comisión bancaria", None),
    "currency_sale": ("Venta de divisas", None),
    "unclassified": ("(sin clasificar)", None),
}

# Sheet 2 acquirer-section header labels (spec §5.4 says "mostrar como
# 'OXXOPay' en reporte"). Used only by sheet_por_adquirente.py.
SHEET2_ACQUIRER_LABEL: dict[str, str] = {
    "kushki": "KUSHKI",
    "bitso": "BITSO",
    "oxxopay": "OXXOPay (vía Pagsmile)",
    "stp": "STP",
    "unlimit": "UNLIMIT",
}


def display_label(db_classification: str) -> tuple[str, str | None]:
    """Map a DB classification code to (display_label, acquirer_code)."""
    return DB_TO_DISPLAY.get(db_classification, (db_classification, None))


def classification_font(label: str) -> Font:
    """Pick the right colored font for a classification label."""
    color = CLASSIFICATION_COLOR.get(label, C.BODY)
    return font(size=9, color=color)
