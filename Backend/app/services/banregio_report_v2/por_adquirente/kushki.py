"""
KUSHKI section of Sheet 2 (spec §4.2.2).

Layout:
  Row N:    section banner '  KUSHKI  —  N depósitos  —  $X MXN  …'
  Rows:     SPEI list (3 cols: Fecha, Descripción, Monto Recibido)
  Blank
  Row:      'Cuadre Banco vs Settlement Report'
  Row:      header (Concepto | Monto | Estatus)
  5 rows:   cuadre values (banco / SR / diff / transit / total devengado)
  Blank
  Row:      'Desglose por comercio'
  Row:      18-col header
  N rows:   per-merchant breakdown
  Row:      TOTAL KUSHKI

Data sources:
  - Banregio movements where classification = kushki_acquirer  →  SPEI list
  - kushki_intransit.classify_rows(KushkiResult.daily_summary)  →  cuadre rows
  - KushkiResult.merchant_detail                                →  desglose 'Kushki' cols
  - FEES file totals_by_merchant_acquirer                       →  desglose 'Tonder' cols
"""
from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import styles as st
from . import _common as cm


SPANISH_MONTHS_LOWER = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def write(
    ws: Worksheet,
    start_row: int,
    *,
    banregio_movements: list[dict],
    classifications_by_idx: dict,
    kushki_merchant_detail: list[dict],
    intransit_summary: dict,
    fees_by_merchant_acquirer: dict[str, dict],
    period_year: int,
    period_month: int,
) -> tuple[int, dict]:
    """Build the Kushki section.

    Returns (next_free_row, stats) where stats is used by the resumen.
    """
    month_name = SPANISH_MONTHS_LOWER[period_month - 1]

    # ── 1) Filter Banregio movements to Kushki acquirer ──────────────
    kushki_movs = []
    for idx, mov in enumerate(banregio_movements):
        cls = classifications_by_idx.get(idx)
        if cls and cls.classification == "kushki_acquirer":
            kushki_movs.append((idx, mov))

    n_deposits = len(kushki_movs)
    total_banco = sum(cm.to_float(mov.get("credit")) for _, mov in kushki_movs)

    # ── 2) Section banner ────────────────────────────────────────────
    banner = (
        f"  KUSHKI  —  {n_deposits} depósitos  —  ${total_banco:,.2f} MXN  "
        f"|  Criterio caja: solo depósitos recibidos en {month_name}"
    )
    row = cm.write_section_header(ws, start_row, banner, st.Fill.SECTION_KUSHKI)

    # ── 3) SPEI list ─────────────────────────────────────────────────
    spei_rows = [
        (mov.get("date"), mov.get("description"), cm.to_float(mov.get("credit")))
        for _, mov in kushki_movs
    ]
    row = cm.write_spei_list(ws, row, ("Fecha", "Descripción", "Monto Recibido"), spei_rows)

    # ── 4) Cuadre Banco vs SR ───────────────────────────────────────
    cm._set(ws, f"A{row}", "Cuadre Banco vs Settlement Report",
            fontobj=st.SECTION_FONT_BLUE, alignment=st.LEFT)
    row += 1

    # column headers for the cuadre table
    cm._set(ws, f"A{row}", "Concepto", fontobj=st.COL_HEADER_FONT, alignment=st.LEFT)
    cm._set(ws, f"B{row}", "Monto", fontobj=st.COL_HEADER_FONT, alignment=st.RIGHT)
    cm._set(ws, f"C{row}", "Estatus", fontobj=st.COL_HEADER_FONT, alignment=st.LEFT)
    row += 1

    sr_intra = cm.to_float(intransit_summary.get("intra_month_total"))
    sr_transit = cm.to_float(intransit_summary.get("transit_total"))
    diff_banco_vs_sr = sr_intra - total_banco

    cuadre_rows = [
        (f"SPEI recibidos en Banregio ({month_name})", total_banco,
         "✅ Cuadrado" if abs(diff_banco_vs_sr) <= 0.01 else f"Diff ${diff_banco_vs_sr:,.2f}"),
        (f"Suma SR Kushki fecha_pago {month_name}", sr_intra,
         "✅ Cuadrado" if abs(diff_banco_vs_sr) <= 0.01 else f"Diff ${diff_banco_vs_sr:,.2f}"),
        ("Diferencia banco vs SR", diff_banco_vs_sr,
         f"✅ ${abs(diff_banco_vs_sr):,.2f}" if abs(diff_banco_vs_sr) <= 0.01
         else f"⚠ ${diff_banco_vs_sr:,.2f}"),
        (
            "Depósito en tránsito (1-abr, txns 31-mar)" if period_month != 12
            else "Depósito en tránsito (1-ene, txns 31-dic)",
            sr_transit,
            "⚠️ Excluido — criterio caja" if sr_transit > 0 else "—"
        ),
        (
            f"Total SR Kushki ({month_name[:3]} + tránsito)",
            sr_intra + sr_transit,
            "Referencia devengado",
        ),
    ]
    for concept, amount, status in cuadre_rows:
        cm._set(ws, f"A{row}", concept, fontobj=st.BODY_FONT, alignment=st.LEFT)
        cm._set(ws, f"B{row}", amount, fontobj=st.BODY_FONT,
                alignment=st.RIGHT, number_format=st.FMT_MXN)
        cm._set(ws, f"C{row}", status,
                fontobj=st.font(size=10, color=st.C.GREEN_DARK if "✅" in status
                               else st.C.ORANGE if "⚠" in status else st.C.SUBTITLE),
                alignment=st.LEFT)
        row += 1
    row += 1  # blank

    # ── 5) Desglose por comercio (18 cols) ──────────────────────────
    cm._set(ws, f"A{row}", "Desglose por comercio",
            fontobj=st.SECTION_FONT_BLUE, alignment=st.LEFT)
    row += 1

    desglose_headers = [
        "Comercio", "# Txns", "Monto Bruto", "Bruto Ajustes",
        "Com. Kushki", "IVA Kushki", "Com. K+IVA", "RR Retenido",
        "Devolución", "Contracargo", "Cancelación",
        "Other Fees", "Ajustes",
        "RR Liberado", "Depósito Neto",
        "Com. Tonder s/IVA", "IVA (16%)", "Com. Tonder c/IVA",
    ]
    for i, label in enumerate(desglose_headers):
        col = get_column_letter(i + 1)
        cm._set(ws, f"{col}{row}", label,
                fontobj=st.COL_HEADER_FONT,
                fillobj=st.COL_HEADER_FILL,
                alignment=st.LEFT if i == 0 else st.RIGHT)
    row += 1

    # Per-merchant rows — sort by net_deposit DESC then merchant_name ASC,
    # but pin TONDER (special row) and any zero-net merchants to the bottom
    # so the layout matches the gold example (active merchants first,
    # housekeeping rows last).
    def _sort_key(m):
        name = str(m.get("merchant_name", "")).strip()
        net = cm.to_float(m.get("net_deposit"))
        is_tonder = name.upper() == "TONDER"
        is_zero = abs(net) < 0.01
        # tier: 0=active, 1=zero-net normal merchant, 2=TONDER (always last)
        tier = 2 if is_tonder else (1 if is_zero else 0)
        return (tier, -net, name)

    merchant_rows = sorted(kushki_merchant_detail or [], key=_sort_key)

    # Aggregate totals as we write
    totals = {k: 0.0 for k in (
        "tx_count", "gross_amount", "adjustments",
        "kushki_commission", "iva_kushki_commission", "commission",
        "rolling_reserve", "refund", "chargeback", "void",
        "manual_adj", "rr_released", "net_deposit",
        "tonder_fee", "tonder_iva", "tonder_fee_iva",
    )}

    # TONDER is the only contributor to Other Fees (col L) and RR Liberado
    # (col N). Per the gold example's data shape:
    #   L = TONDER.manual_adj   (Kushki's "Other Fees" / dispute fees)
    #   N = TONDER.rr_released  (Rolling Reserve liberado)
    # Both appear on the TONDER merchant row AND the TOTAL row (since
    # TONDER is the sole contributor, the column total = TONDER's value).
    tonder_other_fees_total = 0.0
    tonder_rr_liberado_total = 0.0

    for m in merchant_rows:
        merchant = str(m.get("merchant_name", "")).strip()
        if not merchant:
            continue
        is_tonder = merchant.upper() == "TONDER"

        # Tonder fee values from FEES file — fuzzy merchant match. Note
        # that this returns {} when no match, leaving fees as 0 — the
        # cross-source merchant name mapping is FinOps-driven and
        # imperfect by spec design (see §2.3 note).
        fees_row = cm.fees_lookup_for_merchant(
            fees_by_merchant_acquirer, merchant, "kushki",
        )
        tonder_fee_siva = cm.to_float(fees_row.get("fee_siva"))
        tonder_iva = cm.to_float(fees_row.get("iva"))
        tonder_fee_civa = cm.to_float(fees_row.get("fee_civa"))

        if is_tonder:
            other_fees_cell = cm.to_float(m.get("manual_adj"))
            rr_liberado_cell = cm.to_float(m.get("rr_released"))
            tonder_other_fees_total += other_fees_cell
            tonder_rr_liberado_total += rr_liberado_cell
        else:
            # Non-TONDER merchants: per-merchant rr_released is 0 in
            # practice (Kushki only releases RR via the TONDER row), but
            # we still pass through the value if present.
            other_fees_cell = 0
            rr_liberado_cell = cm.to_float(m.get("rr_released"))
            tonder_rr_liberado_total += rr_liberado_cell

        values = [
            merchant,
            int(cm.to_float(m.get("tx_count"))),
            cm.to_float(m.get("gross_amount")),
            cm.to_float(m.get("adjustments")),
            cm.to_float(m.get("kushki_commission")),
            cm.to_float(m.get("iva_kushki_commission")),
            cm.to_float(m.get("commission")),  # K + IVA
            cm.to_float(m.get("rolling_reserve")),
            cm.to_float(m.get("refund")),
            cm.to_float(m.get("chargeback")),
            cm.to_float(m.get("void")),
            other_fees_cell,
            0,  # Ajustes (FinOps-driven; placeholder for now)
            rr_liberado_cell,
            cm.to_float(m.get("net_deposit")),
            tonder_fee_siva,
            tonder_iva,
            tonder_fee_civa,
        ]

        for i, val in enumerate(values):
            col = get_column_letter(i + 1)
            if i == 0:  # merchant name
                cm._set(ws, f"{col}{row}", val, fontobj=st.BODY_FONT, alignment=st.LEFT)
            elif i == 1:  # # Txns
                cm._set(ws, f"{col}{row}", val, fontobj=st.BODY_FONT,
                        alignment=st.RIGHT, number_format=st.FMT_INT)
            elif i == 14:  # Depósito Neto — bold blue
                cm._set(ws, f"{col}{row}", val, fontobj=st.SECTION_FONT_BLUE,
                        alignment=st.RIGHT, number_format=st.FMT_MXN)
            else:
                font = st.BODY_RED if isinstance(val, (int, float)) and val < 0 else st.BODY_FONT
                cm._set(ws, f"{col}{row}", val, fontobj=font,
                        alignment=st.RIGHT, number_format=st.FMT_MXN)

        # Accumulate other-column totals (L and N tracked separately above)
        totals["tx_count"] += cm.to_float(m.get("tx_count"))
        totals["gross_amount"] += cm.to_float(m.get("gross_amount"))
        totals["adjustments"] += cm.to_float(m.get("adjustments"))
        totals["kushki_commission"] += cm.to_float(m.get("kushki_commission"))
        totals["iva_kushki_commission"] += cm.to_float(m.get("iva_kushki_commission"))
        totals["commission"] += cm.to_float(m.get("commission"))
        totals["rolling_reserve"] += cm.to_float(m.get("rolling_reserve"))
        totals["refund"] += cm.to_float(m.get("refund"))
        totals["chargeback"] += cm.to_float(m.get("chargeback"))
        totals["void"] += cm.to_float(m.get("void"))
        totals["net_deposit"] += cm.to_float(m.get("net_deposit"))
        totals["tonder_fee"] += tonder_fee_siva
        totals["tonder_iva"] += tonder_iva
        totals["tonder_fee_iva"] += tonder_fee_civa

        row += 1

    # ── TOTAL KUSHKI row ─────────────────────────────────────────────
    # L (Other Fees) = TONDER's rr_released only (cargos adicionales Kushki)
    # N (RR Liberado) = everyone else's rr_released
    total_values = [
        "TOTAL KUSHKI",
        int(totals["tx_count"]),
        round(totals["gross_amount"], 2),
        round(totals["adjustments"], 2),
        round(totals["kushki_commission"], 2),
        round(totals["iva_kushki_commission"], 2),
        round(totals["commission"], 2),
        round(totals["rolling_reserve"], 2),
        round(totals["refund"], 2),
        round(totals["chargeback"], 2),
        round(totals["void"], 2),
        round(tonder_other_fees_total, 2),    # L — Other Fees (TONDER.manual_adj)
        round(0.0, 2),                        # M — Ajustes (FinOps placeholder)
        round(tonder_rr_liberado_total, 2),   # N — RR Liberado (TONDER.rr_released)
        round(totals["net_deposit"], 2),
        round(totals["tonder_fee"], 2),
        round(totals["tonder_iva"], 2),
        round(totals["tonder_fee_iva"], 2),
    ]
    for i, val in enumerate(total_values):
        col = get_column_letter(i + 1)
        if i == 0:
            cm._set(ws, f"{col}{row}", val, fontobj=st.BODY_BOLD,
                    fillobj=st.fill(st.Fill.TOTAL), alignment=st.LEFT)
        elif i == 1:
            cm._set(ws, f"{col}{row}", val, fontobj=st.BODY_BOLD,
                    fillobj=st.fill(st.Fill.TOTAL),
                    alignment=st.RIGHT, number_format=st.FMT_INT)
        else:
            cm._set(ws, f"{col}{row}", val, fontobj=st.BODY_BOLD,
                    fillobj=st.fill(st.Fill.TOTAL),
                    alignment=st.RIGHT, number_format=st.FMT_MXN)
    row += 1

    stats = {
        "n_deposits": n_deposits,
        "total_banco": round(total_banco, 2),
        "total_sr_intra": round(sr_intra, 2),
        "total_sr_transit": round(sr_transit, 2),
        "diferencia": round(diff_banco_vs_sr, 2),
    }
    return row + 1, stats  # +1 blank
