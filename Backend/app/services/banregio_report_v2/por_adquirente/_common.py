"""
Shared row-writer primitives for Sheet 2 sections.

Each acquirer section gets its own module (kushki.py, bitso.py, …) but
they all share the same header/SPEI-list/cuadre patterns from spec §4.2.
This module owns those patterns so the section files stay focused on
data-mapping logic.

All functions return the next free row number so sections compose with:
    row = write_section_header(ws, row, "  KUSHKI  —  …")
    row = write_spei_list(ws, row, ...)
    row = write_cuadre_table(ws, row, ...)
    return row  # caller continues here
"""
from __future__ import annotations

from typing import Any, Sequence

from openpyxl.worksheet.worksheet import Worksheet

from .. import styles as st


def _set(ws: Worksheet, coord: str, value: Any, *,
         fontobj=None, fillobj=None, alignment=None, number_format=None) -> None:
    cell = ws[coord]
    cell.value = value
    if fontobj is not None:
        cell.font = fontobj
    if fillobj is not None:
        cell.fill = fillobj
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def to_float(v: Any) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _norm_merchant_key(name: str) -> str:
    """Normalize a merchant name for cross-source matching.

    Uppercases and strips all whitespace so 'AFUN'/'Afun Mexico' /
    'BC GAME'/'BCGAME' etc collapse to a comparable form.
    """
    return "".join(str(name or "").upper().split())


def fees_lookup_for_merchant(
    fees_by_merchant_acquirer: dict[str, dict],
    sr_merchant_name: str,
    acquirer: str,
) -> dict:
    """Find the FEES row for a given (SR merchant, acquirer) pair.

    SR Kushki merchant names ('AFUN', 'IDEM CLUB') often differ from FEES
    file merchant names ('Afun Mexico', 'Idem Club'). The naming drift is
    a FinOps reality and there's no canonical mapping yet. We try, in
    order:
      1. Exact match on 'merchant|acquirer'
      2. Normalized exact match (uppercase + no spaces)
      3. UNIQUE normalized prefix match — only if exactly one FEES row
         normalizes to a string that's a prefix of (or prefixed by) the
         SR name. If multiple candidates match, we return {} to avoid
         silently picking the wrong row.

    Returns {} when no safe match exists; caller fills 0 and FinOps
    fixes via a future merchant-mapping config table.
    """
    # 1) exact
    exact = fees_by_merchant_acquirer.get(f"{sr_merchant_name}|{acquirer}", {})
    if exact:
        return exact

    sr_norm = _norm_merchant_key(sr_merchant_name)
    if not sr_norm:
        return {}

    # Collect FEES rows for this acquirer keyed by normalized merchant
    by_norm: dict[str, list[tuple[str, dict]]] = {}
    for key, value in fees_by_merchant_acquirer.items():
        if not key.endswith(f"|{acquirer}"):
            continue
        fees_merchant = key.rsplit("|", 1)[0]
        fees_norm = _norm_merchant_key(fees_merchant)
        if not fees_norm:
            continue
        by_norm.setdefault(fees_norm, []).append((fees_merchant, value))

    # 2) exact normalized match
    if sr_norm in by_norm and len(by_norm[sr_norm]) == 1:
        return by_norm[sr_norm][0][1]

    # 3) unique prefix match — only if exactly one candidate
    candidates = [
        (fnorm, rows[0][1])
        for fnorm, rows in by_norm.items()
        if (fnorm.startswith(sr_norm) or sr_norm.startswith(fnorm)) and len(rows) == 1
    ]
    if len(candidates) == 1:
        return candidates[0][1]

    return {}


def write_global_header(ws: Worksheet, period_label: str) -> int:
    """Rows 1–2 — global header. Returns next free row (4)."""
    _set(ws, "A1", "TrueBook — Desglose por Adquirente",
         fontobj=st.TITLE_FONT, alignment=st.LEFT)
    _set(ws, "A2",
         f"{period_label}  |  Banregio  |  Cierre {period_label.title()}  |  Criterio: Caja",
         fontobj=st.SUBTITLE_FONT, alignment=st.LEFT)
    return 4


def write_section_header(ws: Worksheet, row: int, text: str,
                         fill_token: str = st.Fill.SECTION_KUSHKI) -> int:
    """One-row banner like '  KUSHKI  —  21 depósitos  —  $73M …'.

    Returns row + 1.
    """
    _set(ws, f"A{row}", text,
         fontobj=st.SECTION_FONT_BLUE, fillobj=st.fill(fill_token),
         alignment=st.LEFT)
    # Continue the fill across cols B-S so the banner reads as one band
    for col_idx in range(2, 20):
        from openpyxl.utils import get_column_letter
        c = get_column_letter(col_idx)
        cell = ws[f"{c}{row}"]
        if cell.value is None:
            cell.fill = st.fill(fill_token)
    return row + 1


def write_spei_list(ws: Worksheet, row: int,
                    headers: Sequence[str],
                    rows: list[tuple]) -> int:
    """SPEI / movement list block.

    Args:
        headers: e.g. ("Fecha", "Descripción", "Monto Recibido")
        rows: list of tuples matching the header columns.

    Returns: next free row (one blank after the list).
    """
    # Column header row
    for i, label in enumerate(headers):
        from openpyxl.utils import get_column_letter
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", label,
             fontobj=st.COL_HEADER_FONT, alignment=st.LEFT)
    row += 1

    # Data rows
    for tup in rows:
        for i, val in enumerate(tup):
            from openpyxl.utils import get_column_letter
            col = get_column_letter(i + 1)
            if val is None:
                continue
            if isinstance(val, (int, float)):
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_FONT, alignment=st.RIGHT,
                     number_format=st.FMT_MXN)
            else:
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_FONT, alignment=st.LEFT)
        row += 1

    return row + 1  # +1 blank


def write_table(ws: Worksheet, row: int,
                headers: Sequence[str],
                rows: list[tuple],
                *,
                total_row: tuple | None = None,
                section_fill: str = st.Fill.WHITE) -> int:
    """Generic table block with optional TOTAL row.

    Args:
        headers: column labels (col A onwards)
        rows: data tuples
        total_row: optional last row in bold + filled

    Returns: next free row.
    """
    from openpyxl.utils import get_column_letter

    # Header row
    for i, label in enumerate(headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", label,
             fontobj=st.COL_HEADER_FONT,
             fillobj=st.COL_HEADER_FILL,
             alignment=st.LEFT)
    row += 1

    # Data rows
    for tup in rows:
        for i, val in enumerate(tup):
            col = get_column_letter(i + 1)
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)):
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_FONT, alignment=st.RIGHT,
                     number_format=st.FMT_MXN)
            else:
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_FONT, alignment=st.LEFT)
        row += 1

    # TOTAL row (optional)
    if total_row is not None:
        for i, val in enumerate(total_row):
            col = get_column_letter(i + 1)
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)):
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_BOLD,
                     fillobj=st.fill(st.Fill.TOTAL),
                     alignment=st.RIGHT, number_format=st.FMT_MXN)
            else:
                _set(ws, f"{col}{row}", val,
                     fontobj=st.BODY_BOLD,
                     fillobj=st.fill(st.Fill.TOTAL),
                     alignment=st.LEFT)
        row += 1

    return row + 1  # blank separator


def status_for_diff(diff: float, *, threshold_minor: float = 500.0,
                    threshold_major: float = 500.0,
                    has_pending_flag: bool = False,
                    is_rr_release: bool = False,
                    is_in_transit: bool = False) -> tuple[str, str]:
    """Compute (text, color_hex) per spec §4.2.8 status logic.

    Args:
        diff: signed difference (FEES/SR neto − recibido_banco)
        threshold_minor: under this absolute diff → "Diferencia menor"
        threshold_major: above this → CRITICAL
        has_pending_flag: merchant flagged in pending_transfer_merchants
        is_rr_release: Unlimit-style sections where diff is always 0
        is_in_transit: cash-basis exclusion
    """
    abs_diff = abs(diff)
    if is_in_transit:
        return ("⚠️ Excluido — criterio caja", st.C.ORANGE)
    if is_rr_release and abs_diff <= 0.01:
        return ("✅ RR liberado", st.C.GREEN_DARK)
    if abs_diff <= 0.01:
        return ("✅ Cuadrado", st.C.GREEN_DARK)
    if has_pending_flag:
        return ("⚠️ Pendiente de transferir", st.C.RED)
    if abs_diff <= threshold_minor:
        return ("🔍 Diferencia menor", st.C.ORANGE)
    if abs_diff > threshold_major:
        return ("❓ En investigación", st.C.ORANGE)
    return (f"Diff ${abs_diff:,.2f}", st.C.ORANGE)
