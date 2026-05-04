"""Resumen Consolidado section of Sheet 2 (spec §4.2.7).

Final 7-row table at the bottom of Sheet 2:
  Adquirente | Neto FEES/SR | Recibido Banco | Diferencia | Estatus

Rows: Kushki, Bitso-CampoBet, Bitso-Artilu, OXXOPay, STP, Unlimit, TOTAL.

Takes the stats dict returned by each section above and renders the
single source of truth for FinOps sign-off.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import styles as st
from . import _common as cm


def write(
    ws: Worksheet,
    start_row: int,
    *,
    kushki_stats: dict,
    bitso_stats: dict,
    oxxopay_stats: dict,
    stp_stats: dict,
    unlimit_stats: dict,
    umbral_minor: float,
    umbral_major: float,
) -> tuple[int, dict]:
    # Banner
    cm._set(ws, f"A{start_row}", "  RESUMEN CONSOLIDADO — TODOS LOS ADQUIRENTES",
            fontobj=st.font(size=11, color=st.C.WHITE, bold=True),
            fillobj=st.fill(st.Fill.SECTION_RESUMEN),
            alignment=st.LEFT)
    # Continue dark fill across cols
    for col_idx in range(2, 6):
        col = get_column_letter(col_idx)
        ws[f"{col}{start_row}"].fill = st.fill(st.Fill.SECTION_RESUMEN)
    row = start_row + 1

    # Header
    headers = ["Adquirente", "Neto FEES / SR", "Recibido Banco", "Diferencia", "Estatus"]
    for i, label in enumerate(headers):
        col = get_column_letter(i + 1)
        cm._set(ws, f"{col}{row}", label,
                fontobj=st.COL_HEADER_FONT, fillobj=st.COL_HEADER_FILL,
                alignment=st.LEFT if i in (0, 4) else st.RIGHT)
    row += 1

    def _status(diff: float, *, is_rr: bool = False, is_pending: bool = False) -> tuple[str, str]:
        abs_d = abs(diff)
        if is_rr and abs_d <= 0.01:
            return ("✅ RR liberado", st.C.GREEN_DARK)
        if abs_d <= 0.01:
            return ("✅ Cuadrado", st.C.GREEN_DARK)
        if is_pending:
            return ("⚠️ Pendiente de transferir", st.C.RED)
        if abs_d > umbral_major:
            return ("❓ En investigación", st.C.ORANGE)
        if abs_d <= umbral_minor:
            return ("🔍 Diferencia menor", st.C.ORANGE)
        return (f"⚠ ${diff:,.2f}", st.C.ORANGE)

    # Build rows
    rows_data: list[tuple] = []

    # Kushki
    rows_data.append((
        "Kushki",
        kushki_stats["total_sr_intra"],
        kushki_stats["total_banco"],
        kushki_stats["diferencia"],
        _status(kushki_stats["diferencia"]),
    ))

    # Bitso — split into CampoBet and Artilu MX
    rows_data.append((
        "Bitso — CampoBet",
        bitso_stats["campobet_neto_liquidar"],
        bitso_stats["campobet_recibido"],
        bitso_stats["campobet_diferencia"],
        _status(bitso_stats["campobet_diferencia"]),
    ))
    rows_data.append((
        "Bitso — Artilu MX",
        bitso_stats["artilu_neto_liquidar"],
        bitso_stats["artilu_recibido"],
        bitso_stats["artilu_diferencia"],
        _status(bitso_stats["artilu_diferencia"], is_pending=True),
    ))

    # OXXOPay
    rows_data.append((
        "OXXOPay (Pagsmile)",
        oxxopay_stats["neto_fees"],
        oxxopay_stats["total_banco"],
        oxxopay_stats["diferencia"],
        _status(oxxopay_stats["diferencia"]),
    ))

    # STP
    rows_data.append((
        "STP",
        stp_stats["neto_fees"],
        stp_stats["total_banco"],
        stp_stats["diferencia"],
        _status(stp_stats["diferencia"], is_pending=stp_stats.get("has_pending", False)),
    ))

    # Unlimit (RR-style)
    rows_data.append((
        "Unlimit",
        unlimit_stats["total_banco"],
        unlimit_stats["total_banco"],
        0.0,
        _status(0.0, is_rr=True),
    ))

    # Write rows
    for label, neto, banco, diff, (status_text, status_color) in rows_data:
        cm._set(ws, f"A{row}", label, fontobj=st.BODY_FONT, alignment=st.LEFT)
        cm._set(ws, f"B{row}", neto, fontobj=st.BODY_FONT,
                alignment=st.RIGHT, number_format=st.FMT_MXN)
        cm._set(ws, f"C{row}", banco, fontobj=st.BODY_FONT,
                alignment=st.RIGHT, number_format=st.FMT_MXN)
        diff_color = st.C.GREEN_DARK if abs(diff) <= 0.01 else \
                     st.C.RED if diff > 0 else st.C.ORANGE
        cm._set(ws, f"D{row}", diff,
                fontobj=st.font(size=10, color=diff_color, bold=abs(diff) > 0.01),
                alignment=st.RIGHT, number_format=st.FMT_MXN)
        cm._set(ws, f"E{row}", status_text,
                fontobj=st.font(size=9, color=status_color, bold=True),
                alignment=st.LEFT)
        row += 1

    # TOTAL row
    total_neto = sum(r[1] for r in rows_data)
    total_banco = sum(r[2] for r in rows_data)
    total_diff = sum(r[3] for r in rows_data)
    cm._set(ws, f"A{row}", "TOTAL",
            fontobj=st.BODY_BOLD, fillobj=st.fill(st.Fill.TOTAL), alignment=st.LEFT)
    cm._set(ws, f"B{row}", round(total_neto, 2),
            fontobj=st.BODY_BOLD, fillobj=st.fill(st.Fill.TOTAL),
            alignment=st.RIGHT, number_format=st.FMT_MXN)
    cm._set(ws, f"C{row}", round(total_banco, 2),
            fontobj=st.BODY_BOLD, fillobj=st.fill(st.Fill.TOTAL),
            alignment=st.RIGHT, number_format=st.FMT_MXN)
    cm._set(ws, f"D{row}", round(total_diff, 2),
            fontobj=st.BODY_BOLD, fillobj=st.fill(st.Fill.TOTAL),
            alignment=st.RIGHT, number_format=st.FMT_MXN)
    row += 1

    return row, {
        "total_neto": round(total_neto, 2),
        "total_banco": round(total_banco, 2),
        "total_diferencia": round(total_diff, 2),
    }
