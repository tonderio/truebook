"""
Builder for Sheet 3 — `Alertas` (spec v2 §4.3).

Two stacked tables:
  1. Pending Banregio movements (anything with classification=unclassified)
  2. System alerts (TIMING_CAJA, PENDING_TRANSFER, etc.)

The alert rows themselves come from `alert_generator.generate(...)` —
this module is presentation-only.
"""
from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.process import AccountingProcess
from app.models.result import BanregioResult
from app.models.classification import BanregioMovementClassification

from . import styles as st
from . import alert_generator
from app.services import kushki_intransit


SPANISH_MONTHS = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

COLUMN_WIDTHS = {
    "A": 20.0,
    "B": 25.0,
    "C": 45.0,
    "D": 55.0,
}


# Level → (font_color, fill_color) tokens from the gold inspection
LEVEL_STYLE = {
    "INFO": (st.C.BLUE_BOLD, st.Fill.COL_HEADER),
    "WARNING": (st.C.ORANGE, st.Fill.ALERT_YELLOW),
    "INVESTIGATION": (st.C.ORANGE, st.Fill.ALERT_ORANGE),
    "CRITICAL": (st.C.RED, st.Fill.ALERT_RED),
}


def _set(ws: Worksheet, coord: str, value: Any, *,
         fontobj=None, fillobj=None, alignment=None, number_format=None) -> None:
    cell = ws[coord]
    cell.value = value
    if fontobj is not None:
        cell.font = fontobj
    if fillobj is not None:
        cell.fill = fillobj
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def _write_pending_section(
    ws: Worksheet,
    start_row: int,
    pending_movements: list[tuple[int, dict]],
) -> int:
    """Section header + 6-col table of pending movements (or '(ninguno)')."""
    # Section banner
    _set(ws, f"A{start_row}", "MOVIMIENTOS PENDIENTES DE RECONCILIAR",
         fontobj=st.font(size=11, color=st.C.TITLE, bold=True),
         alignment=st.LEFT)
    row = start_row + 1

    # Column headers
    headers = ["#", "Fecha", "Descripción", "Cargo", "Abono", "Acción requerida"]
    for i, label in enumerate(headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", label,
             fontobj=st.font(size=10, color=st.C.SUBTITLE, bold=True),
             fillobj=st.fill(st.Fill.COL_HEADER),
             alignment=st.LEFT if i in (0, 1, 2, 5) else st.RIGHT)
    row += 1

    if not pending_movements:
        _set(ws, f"A{row}", "(ninguno)",
             fontobj=st.font(size=10, color=st.C.SUBTITLE),
             alignment=st.LEFT)
        row += 1
    else:
        for n, (idx, mov) in enumerate(pending_movements, start=1):
            _set(ws, f"A{row}", n, fontobj=st.BODY_FONT, alignment=st.LEFT)
            _set(ws, f"B{row}", mov.get("date") or "",
                 fontobj=st.BODY_FONT, alignment=st.LEFT)
            _set(ws, f"C{row}", mov.get("description") or "",
                 fontobj=st.BODY_FONT, alignment=st.LEFT)
            debit = float(mov.get("debit") or 0)
            credit = float(mov.get("credit") or 0)
            if debit > 0:
                _set(ws, f"D{row}", debit,
                     fontobj=st.BODY_RED, alignment=st.RIGHT,
                     number_format=st.FMT_MXN)
            if credit > 0:
                _set(ws, f"E{row}", credit,
                     fontobj=st.BODY_FONT, alignment=st.RIGHT,
                     number_format=st.FMT_MXN)
            _set(ws, f"F{row}", "Clasificar manualmente",
                 fontobj=st.font(size=10, color=st.C.ORANGE),
                 alignment=st.LEFT)
            row += 1

    return row + 1  # +1 blank


def _write_alerts_section(
    ws: Worksheet,
    start_row: int,
    alerts: list[dict],
) -> int:
    """Section header + 4-col table of system alerts."""
    # Section banner
    _set(ws, f"A{start_row}", "ALERTAS DEL SISTEMA",
         fontobj=st.font(size=11, color=st.C.TITLE, bold=True),
         alignment=st.LEFT)
    row = start_row + 1

    # Headers
    headers = ["Nivel", "Tipo", "Título", "Mensaje"]
    for i, label in enumerate(headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", label,
             fontobj=st.font(size=10, color=st.C.SUBTITLE, bold=True),
             fillobj=st.fill(st.Fill.COL_HEADER),
             alignment=st.LEFT)
    row += 1

    if not alerts:
        _set(ws, f"A{row}", "(sin alertas)",
             fontobj=st.font(size=10, color=st.C.SUBTITLE),
             alignment=st.LEFT)
        row += 1
        return row + 1

    for alert in alerts:
        level = alert["level"]
        font_color, fill_color = LEVEL_STYLE.get(
            level, (st.C.SUBTITLE, st.Fill.WHITE)
        )
        fillobj = st.fill(fill_color)

        # Col A — Nivel (level label, bold colored)
        _set(ws, f"A{row}", level,
             fontobj=st.font(size=10, color=font_color, bold=True),
             fillobj=fillobj, alignment=st.LEFT)

        # Col B — Tipo (type code)
        _set(ws, f"B{row}", alert["type"],
             fontobj=st.font(size=10, color=st.C.BODY),
             fillobj=fillobj, alignment=st.LEFT)

        # Col C — Título (bold)
        _set(ws, f"C{row}", alert["title"],
             fontobj=st.font(size=10, color=st.C.TITLE, bold=True),
             fillobj=fillobj, alignment=st.LEFT)

        # Col D — Mensaje
        _set(ws, f"D{row}", alert["message"],
             fontobj=st.font(size=10, color=st.C.BODY),
             fillobj=fillobj, alignment=st.WRAP)
        row += 1

    return row


def build(
    ws: Worksheet,
    db: Session,
    process: AccountingProcess,
    *,
    sheet2_stats: dict,
) -> dict:
    """Build Sheet 3 for `process`. `sheet2_stats` is the dict returned
    by `sheet_por_adquirente.build()`.

    Returns a stats dict with alert counts.
    """
    # Column widths
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # ── Pending movements ───────────────────────────────────────────
    br = db.query(BanregioResult).filter_by(process_id=process.id).first()
    movements = (br.movements if br else []) or []
    classifications = (
        db.query(BanregioMovementClassification)
        .filter_by(process_id=process.id)
        .all()
    )
    by_idx = {c.movement_index: c for c in classifications}

    pending: list[tuple[int, dict]] = []
    for idx, mov in enumerate(movements):
        cls = by_idx.get(idx)
        if not cls or cls.classification in (None, "unclassified", ""):
            pending.append((idx, mov))

    pending_count = len(pending)
    coverage_pct = float(process.coverage_pct or 0)

    # ── In-transit classification (for TIMING_CAJA alert) ───────────
    from app.models.result import KushkiResult
    kr = db.query(KushkiResult).filter_by(process_id=process.id).first()
    intransit = kushki_intransit.classify_rows(
        (kr.daily_summary if kr else []) or [],
        process.period_year, process.period_month,
    )

    # ── Generate alerts ─────────────────────────────────────────────
    alerts = alert_generator.generate(
        db, process,
        sheet2_stats=sheet2_stats,
        coverage_pct=coverage_pct,
        intransit_classification=intransit,
    )

    # ── Header (rows 1-2) ───────────────────────────────────────────
    month_name = SPANISH_MONTHS[process.period_month - 1]
    year = process.period_year
    _set(ws, "A1", "TrueBook — Alertas de Reconciliación",
         fontobj=st.TITLE_FONT, alignment=st.LEFT)
    _set(ws, "A2",
         f"{month_name} {year}  |  Cobertura: {coverage_pct:.1f}%  |  "
         f"{pending_count} movimientos pendientes",
         fontobj=st.SUBTITLE_FONT, alignment=st.LEFT)

    # ── Section 1 — Pending movements (starts at row 4) ─────────────
    next_row = _write_pending_section(ws, start_row=4, pending_movements=pending)

    # ── Section 2 — System alerts ───────────────────────────────────
    next_row = _write_alerts_section(ws, start_row=next_row, alerts=alerts)

    return {
        "pending_count": pending_count,
        "alert_count": len(alerts),
        "alerts_by_level": {
            level: sum(1 for a in alerts if a["level"] == level)
            for level in ("CRITICAL", "WARNING", "INVESTIGATION", "INFO")
        },
    }
