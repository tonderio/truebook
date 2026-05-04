"""
Builder for Sheet 1 — `Reconciliación` (spec v2 §4.1).

Layout:
  Rows 1-2: title + subtitle (merged across A:J)
  Row 3:    blank
  Row 4:    KPI labels (A=Movimientos, C=Reconciliados, E=Pendientes, G=Cobertura)
  Row 5:    KPI values
  Row 6:    blank
  Row 7:    column headers (B-J)
  Row 8:    SALDO INICIAL (Opening Balance)
  Rows 9+:  one row per Banregio movement
  Row last: SALDO FINAL (Closing Balance)

Reads:
  - AccountingProcess  (period_year, period_month, name, coverage_pct)
  - BanregioResult.movements
  - BanregioMovementClassification (for label + acquirer + status)

Writes:
  - rows on the supplied openpyxl Worksheet (caller owns the workbook)

This module is pure-presentation: no DB writes, no side effects beyond
the Worksheet it's given.
"""
from __future__ import annotations

from datetime import date as date_t
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.process import AccountingProcess
from app.models.result import BanregioResult
from app.models.classification import BanregioMovementClassification

from . import styles as st


SPANISH_MONTHS = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

COLUMN_WIDTHS = {
    "A": 36.0,   # check
    "B": 12.0,   # fecha
    "C": 35.0,   # descripción
    "D": 12.0,   # cargo
    "E": 17.16,  # abono
    "F": 13.0,   # saldo
    "G": 24.0,   # clasificación
    "H": 12.0,   # adquirente
    "I": 10.0,   # método
    "J": 14.0,   # estado
}


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _set(ws: Worksheet, coord: str, value: Any, *, fontobj=None, fillobj=None,
         alignment=None, number_format: str | None = None) -> None:
    """Convenience: write a cell with optional styling."""
    cell = ws[coord]
    cell.value = value
    if fontobj is not None:
        cell.font = fontobj
    if fillobj is not None:
        cell.fill = fillobj
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def _build_header(ws: Worksheet, process: AccountingProcess, stats: dict) -> None:
    """Rows 1-5: title + KPIs."""
    month_name = SPANISH_MONTHS[process.period_month - 1]
    year = process.period_year

    # Row 1 — title
    ws.merge_cells("A1:J1")
    _set(ws, "A1", "TrueBook — Reconciliación Banregio",
         fontobj=st.TITLE_FONT, alignment=st.LEFT)

    # Row 2 — subtitle
    ws.merge_cells("A2:J2")
    _set(ws, "A2", f"{month_name} {year}  |  Cierre {month_name.title()} {year}",
         fontobj=st.SUBTITLE_FONT, alignment=st.LEFT)

    # Row 4 — KPI labels
    _set(ws, "A4", "Movimientos", fontobj=st.KPI_LABEL_FONT)
    _set(ws, "C4", "Reconciliados", fontobj=st.KPI_LABEL_FONT)
    _set(ws, "E4", "Pendientes", fontobj=st.KPI_LABEL_FONT)
    _set(ws, "G4", "Cobertura", fontobj=st.KPI_LABEL_FONT)

    # Row 5 — KPI values
    _set(ws, "A5", stats["total"], fontobj=st.KPI_VALUE_FONT, number_format=st.FMT_INT)
    _set(ws, "C5", stats["reconciliados"], fontobj=st.KPI_VALUE_GREEN,
         number_format=st.FMT_INT)
    _set(ws, "E5", stats["pendientes"],
         fontobj=st.KPI_VALUE_GRAY if stats["pendientes"] == 0 else st.KPI_VALUE_BLUE,
         number_format=st.FMT_INT)
    _set(ws, "G5", f"{stats['cobertura']:.1f}%", fontobj=st.KPI_VALUE_BLUE)


def _build_column_headers(ws: Worksheet, row: int = 7) -> None:
    """Row 7 — column headers (B-J), styled per spec/example."""
    headers = [
        ("B", "Fecha"),
        ("C", "Descripción"),
        ("D", "Cargo"),
        ("E", "Abono"),
        ("F", "Saldo"),
        ("G", "Clasificación"),
        ("H", "Adquirente"),
        ("I", "Método"),
        ("J", "Estado"),
    ]
    for col, label in headers:
        _set(ws, f"{col}{row}", label,
             fontobj=st.COL_HEADER_FONT, fillobj=st.COL_HEADER_FILL)


def _format_date(s: Any) -> str:
    """Banregio movement dates come pre-formatted as DD/MM/YYYY strings.
    Pass through if already a string; format if it's a date.
    """
    if s is None:
        return ""
    if isinstance(s, date_t):
        return s.strftime("%d/%m/%Y")
    return str(s).strip()


def _write_movement_row(
    ws: Worksheet,
    row: int,
    movement: dict,
    classification_label: str,
    acquirer: str | None,
    method: str,
    status: str,
    saldo: float,
) -> None:
    """Write one movement row at the given row number."""
    white = st.fill(st.Fill.WHITE)

    # Column A — check mark for reconciled
    if status == "Reconciliado":
        _set(ws, f"A{row}", "✓", fontobj=st.CHECK_FONT,
             alignment=st.CENTER, fillobj=white)

    # B — fecha
    _set(ws, f"B{row}", _format_date(movement.get("date")),
         fontobj=st.BODY_FONT, fillobj=white, alignment=st.LEFT)

    # C — descripción
    _set(ws, f"C{row}", movement.get("description") or "",
         fontobj=st.BODY_FONT, fillobj=white, alignment=st.LEFT)

    # D — cargo
    debit = _to_float(movement.get("debit"))
    if debit > 0:
        _set(ws, f"D{row}", debit, fontobj=st.BODY_RED, fillobj=white,
             alignment=st.RIGHT, number_format=st.FMT_MXN)

    # E — abono
    credit = _to_float(movement.get("credit"))
    if credit > 0:
        _set(ws, f"E{row}", credit, fontobj=st.BODY_FONT, fillobj=white,
             alignment=st.RIGHT, number_format=st.FMT_MXN)

    # F — saldo (running)
    _set(ws, f"F{row}", saldo, fontobj=st.BODY_BOLD, fillobj=white,
         alignment=st.RIGHT, number_format=st.FMT_MXN)

    # G — clasificación
    if classification_label:
        _set(ws, f"G{row}", classification_label,
             fontobj=st.classification_font(classification_label),
             fillobj=white, alignment=st.LEFT)

    # H — adquirente
    if acquirer:
        _set(ws, f"H{row}", acquirer,
             fontobj=st.font(size=9, color=st.C.SUBTITLE),
             fillobj=white, alignment=st.LEFT)

    # I — método
    _set(ws, f"I{row}", method, fontobj=st.METHOD_FONT, fillobj=white,
         alignment=st.LEFT)

    # J — estado
    status_font = st.STATUS_RECONCILED_FONT if status == "Reconciliado" else st.STATUS_PENDING_FONT
    _set(ws, f"J{row}", status, fontobj=status_font, fillobj=white, alignment=st.LEFT)


def _write_balance_row(
    ws: Worksheet, row: int, label: str, balance: float, *, opening: bool = False
) -> None:
    """SALDO INICIAL or SALDO FINAL row — bold blue with light fill."""
    fillobj = st.SECTION_FILL_BLUE
    fontobj = st.SECTION_FONT_BLUE
    # Spans columns C (label) and F (value)
    _set(ws, f"C{row}", label, fontobj=fontobj, fillobj=fillobj, alignment=st.LEFT)
    _set(ws, f"F{row}", balance, fontobj=fontobj, fillobj=fillobj,
         alignment=st.RIGHT, number_format=st.FMT_MXN)
    # Fill rest of the row light blue too for visual continuity
    for col in ("A", "B", "D", "E", "G", "H", "I", "J"):
        cell = ws[f"{col}{row}"]
        if cell.value is None:
            cell.fill = fillobj


# ── public entry point ───────────────────────────────────────────────────


def build(ws: Worksheet, db: Session, process: AccountingProcess,
          opening_balance: float = 0.0) -> dict:
    """Populate `ws` with the full Reconciliación sheet for `process`.

    Args:
        ws: target Worksheet (caller owns the workbook).
        db: SQLAlchemy session for reading classifications.
        process: AccountingProcess row (must have .id, .period_year, .period_month).
        opening_balance: SALDO INICIAL value (default 0.00 — matches the
            March example; if you have a prior closing balance from
            previous period, pass it here).

    Returns:
        Stats dict: {total, reconciliados, pendientes, cobertura,
                     opening_balance, closing_balance, total_credits, total_debits}
    """
    # Data fetch ─────────────────────────────────────────────────────
    br = db.query(BanregioResult).filter_by(process_id=process.id).first()
    movements = (br.movements if br else []) or []

    classifications = (
        db.query(BanregioMovementClassification)
        .filter_by(process_id=process.id)
        .all()
    )
    by_idx = {c.movement_index: c for c in classifications}

    # Aggregate stats ────────────────────────────────────────────────
    total = len(movements)
    reconciled_count = 0
    pendientes = 0
    sum_credits = 0.0
    sum_debits = 0.0

    for idx, mov in enumerate(movements):
        sum_credits += _to_float(mov.get("credit"))
        sum_debits += _to_float(mov.get("debit"))
        cls = by_idx.get(idx)
        if cls and cls.classification not in (None, "unclassified", ""):
            reconciled_count += 1
        else:
            pendientes += 1

    cobertura = (reconciled_count / total * 100.0) if total else 0.0
    closing_balance = opening_balance + sum_credits - sum_debits

    stats = {
        "total": total,
        "reconciliados": reconciled_count,
        "pendientes": pendientes,
        "cobertura": cobertura,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "total_credits": round(sum_credits, 2),
        "total_debits": round(sum_debits, 2),
    }

    # Header (rows 1-5) ──────────────────────────────────────────────
    _build_header(ws, process, stats)

    # Column widths ─────────────────────────────────────────────────
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Column headers (row 7) ────────────────────────────────────────
    _build_column_headers(ws, row=7)

    # Saldo inicial (row 8) ─────────────────────────────────────────
    _write_balance_row(ws, 8, "SALDO INICIAL (Opening Balance)",
                       opening_balance, opening=True)

    # Movement rows (row 9 onwards) ─────────────────────────────────
    saldo = opening_balance
    row_num = 9
    for idx, mov in enumerate(movements):
        debit = _to_float(mov.get("debit"))
        credit = _to_float(mov.get("credit"))
        # Round to centavos at every step — running sums of floats
        # accumulate IEEE-754 noise that produces ugly cells like
        # 3913636.000000002. Saldo is in MXN, 2-decimal precision.
        saldo = round(saldo + credit - debit, 2)

        cls = by_idx.get(idx)
        if cls:
            label, _ = st.display_label(cls.classification)
            acquirer = cls.acquirer
            method = cls.classification_method or "auto"
            status = "Reconciliado" if cls.classification not in (None, "unclassified", "") else "Pendiente"
        else:
            label, acquirer, method, status = "", None, "auto", "Pendiente"

        _write_movement_row(
            ws=ws,
            row=row_num,
            movement=mov,
            classification_label=label,
            acquirer=acquirer,
            method=method,
            status=status,
            saldo=saldo,
        )
        row_num += 1

    # Saldo final (last row) ────────────────────────────────────────
    _write_balance_row(ws, row_num, "SALDO FINAL (Closing Balance)",
                       closing_balance, opening=False)

    # Freeze panes below header so column headers stay visible while scrolling
    ws.freeze_panes = "A8"

    return stats
