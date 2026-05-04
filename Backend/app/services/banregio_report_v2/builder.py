"""
Top-level workbook builder for the v2 Banregio Reconciliation Report.

Composes Sheet 1 + Sheet 2 + Sheet 3 in one entry point. Used by both
the CLI script (`scripts/generate_banregio_report.py`) and the HTTP
endpoint (`app/routers/banregio_report.py`) — single source of truth.

Returns the openpyxl Workbook object so callers can save it wherever
they want (disk for CLI, BytesIO for HTTP).
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.models.process import AccountingProcess
from . import sheet_reconciliacion, sheet_por_adquirente, sheet_alertas


SPANISH_MONTHS_FILE = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]


def default_filename(process: AccountingProcess) -> str:
    """The canonical filename FinOps expects per the v2 spec.

    Format: `RECONCILIACION_BANREGIO_{MES}_{AÑO}_v2.xlsx` — uppercase
    Spanish month name, year. Matches the gold example's filename.
    """
    month_name = SPANISH_MONTHS_FILE[process.period_month - 1]
    return f"RECONCILIACION_BANREGIO_{month_name}_{process.period_year}_v2.xlsx"


def build_workbook(
    db: Session,
    process: AccountingProcess,
    *,
    opening_balance: float = 0.0,
) -> tuple[openpyxl.Workbook, dict]:
    """Build the full 3-sheet v2 workbook for a given run.

    Args:
        db: SQLAlchemy session.
        process: AccountingProcess row (period_year, period_month, etc).
        opening_balance: SALDO INICIAL on Sheet 1. Defaults to 0.

    Returns:
        (workbook, stats) where stats merges all three sheets'
        per-sheet stats dicts under keys `sheet1`, `sheet2`, `sheet3`.
    """
    wb = openpyxl.Workbook()

    # Sheet 1 — Reconciliación
    ws1 = wb.active
    ws1.title = "Reconciliación"
    sheet1_stats = sheet_reconciliacion.build(
        ws1, db, process, opening_balance=opening_balance
    )

    # Sheet 2 — Por Adquirente
    ws2 = wb.create_sheet("Por Adquirente")
    sheet2_stats = sheet_por_adquirente.build(ws2, db, process)

    # Sheet 3 — Alertas
    ws3 = wb.create_sheet("Alertas")
    sheet3_stats = sheet_alertas.build(ws3, db, process, sheet2_stats=sheet2_stats)

    return wb, {
        "sheet1": sheet1_stats,
        "sheet2": sheet2_stats,
        "sheet3": sheet3_stats,
    }


def build_to_bytes(
    db: Session,
    process: AccountingProcess,
    *,
    opening_balance: float = 0.0,
) -> tuple[bytes, dict]:
    """Build the workbook and return its serialized .xlsx bytes.

    Convenience for the HTTP endpoint where we stream the file rather
    than persisting to disk.
    """
    wb, stats = build_workbook(db, process, opening_balance=opening_balance)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), stats


def build_to_path(
    db: Session,
    process: AccountingProcess,
    path: Path | str,
    *,
    opening_balance: float = 0.0,
) -> dict:
    """Build the workbook and save it to `path`. Creates parent dirs.

    Used by the CLI and by the HTTP endpoint when also persisting an
    audit copy to `uploads/{process_id}/reports/`.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb, stats = build_workbook(db, process, opening_balance=opening_balance)
    wb.save(path)
    return stats
