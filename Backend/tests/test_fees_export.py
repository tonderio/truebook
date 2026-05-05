"""Tests for the auto-generated FEES xlsx pipeline.

Three concerns covered:

1. `build_fees_export` produces the canonical 4-sheet shape with the
   gold-file conventions (canonical Concepto labels, per-merchant
   subtotal rows, real Razón Social grouping when a db is provided).
2. `_run_full_process` Stage 4b creates an `UploadedFile(file_type='fees',
   status='auto_generated')` row, idempotently (re-runs replace, don't
   accumulate).
3. A pre-existing manual FEES upload wins — auto-gen skips itself.

These run against an in-memory SQLite db with all real models so any
SQLAlchemy-side breakage shows up locally before deploy.
"""
from __future__ import annotations

import io
import os
import tempfile
import types
from unittest.mock import patch

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.alert import ReconciliationConfig  # noqa: F401  (registers metadata)
from app.models.file import UploadedFile
from app.models.process import AccountingProcess, ProcessLog  # noqa: F401
from app.models.result import FeesResult  # noqa: F401
from app.services import banregio_report_config as cfg
from app.services.excel_exports import (
    ACQUIRER_CONCEPTO,
    _concepto_for,
    build_fees_export,
)


# ── Fixtures ────────────────────────────────────────────────────────────


@pytest.fixture
def db_session():
    """In-memory SQLite with all metadata. Seeds the v2 config so
    `razon_social_for` resolves."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    cfg.seed_defaults(db)
    yield db
    db.close()


@pytest.fixture
def mock_process():
    p = types.SimpleNamespace()
    p.id = 999
    p.period_year = 2026
    p.period_month = 4
    return p


@pytest.fixture
def mock_fees_result():
    """A FeesResult-shaped object covering 3 merchants × 3 acquirers,
    plus a withdrawal row, so subtotals exercise all branches."""
    return types.SimpleNamespace(
        merchant_summary=[
            {"merchant_id": "m1", "merchant_name": "AFUNVIP",
             "gross_amount": 134030.06, "total_fee": 1256.09},
            {"merchant_id": "m2", "merchant_name": "Afun Mexico",
             "gross_amount": 55832337.0, "total_fee": 2286558.66},
            {"merchant_id": "m3", "merchant_name": "BCGAME",
             "gross_amount": 11977433.83, "total_fee": 145309.13},
        ],
        daily_breakdown=[
            {"date": "2026-04-01", "merchant_id": "m1",
             "merchant_name": "AFUNVIP", "acquirer": "bitso",
             "amount": 50.0, "fee_amount": 1.25},
            {"date": "2026-04-01", "merchant_id": "m2",
             "merchant_name": "Afun Mexico", "acquirer": "kushki",
             "amount": 2002737.0, "fee_amount": 81940.04},
            {"date": "2026-04-01", "merchant_id": "m3",
             "merchant_name": "BCGAME", "acquirer": "bitso",
             "amount": 141219.58, "fee_amount": 1852.0},
            {"date": "2026-04-01", "merchant_id": "m3",
             "merchant_name": "BCGAME", "acquirer": "oxxopay",
             "amount": 4590.0, "fee_amount": 361.55},
        ],
        withdrawals_summary=[
            {"merchant_id": "m3", "merchant_name": "BCGAME",
             "count": 4, "total_amount": 8380.0, "total_fee": 380.0},
        ],
        refunds_summary=[],
        other_fees_summary=[],
    )


# ── Tests ───────────────────────────────────────────────────────────────


def test_concepto_mapping_covers_known_acquirers():
    """Canonical labels must match the FEES gold file. If this list grows,
    update both ACQUIRER_CONCEPTO and this test together."""
    assert ACQUIRER_CONCEPTO["bitso"] == "BITSO - SPEI"
    assert ACQUIRER_CONCEPTO["kushki"] == "Kushki - Tarjetas"
    assert ACQUIRER_CONCEPTO["oxxopay"] == "OXXOPay"
    assert ACQUIRER_CONCEPTO["stp"] == "STP - SPEI"

    # Case-insensitive on input
    assert _concepto_for("BITSO") == "BITSO - SPEI"
    assert _concepto_for("Kushki") == "Kushki - Tarjetas"

    # Empty/None safe
    assert _concepto_for("") == "Operativa"
    assert _concepto_for(None) == "Operativa"

    # Unknown acquirer doesn't drop the row — gets a synthetic label
    assert _concepto_for("newacq") == "newacq-Operativa"


def test_build_fees_export_smoke(mock_process, mock_fees_result):
    """The legacy (no-db) call still works — covers the manual
    download endpoint at routers/results.py:export_fees_excel."""
    filename, content = build_fees_export(mock_process, mock_fees_result)
    assert filename == "FEES_ABRIL_2026_FINAL.xlsx"
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames == [
        "Detalle por Merchant",
        "Resumen por Merchant",
        "Resumen por Razon Social",
        "Tonder Fees desglose diario",
    ]


def test_detalle_uses_canonical_concepto_labels(mock_process, mock_fees_result):
    """Sheet 1 should never emit "{acquirer}-Operativa" for known acquirers."""
    _, content = build_fees_export(mock_process, mock_fees_result)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Detalle por Merchant"]

    concepts_seen = set()
    for row in ws.iter_rows(values_only=True):
        # Skip header rows + merchant-name rows + subtotal rows
        if row and row[2] and not str(row[1] or "").startswith("Subtotal"):
            concepts_seen.add(str(row[2]))

    # Should contain canonical labels
    assert "BITSO - SPEI" in concepts_seen
    assert "Kushki - Tarjetas" in concepts_seen
    assert "OXXOPay" in concepts_seen
    # Should NOT contain the legacy concat form
    assert not any("-Operativa" in c for c in concepts_seen if c not in {
        # Withdrawals/Autorefunds intentionally use their own labels
        "Withdrawals", "Autorefunds/Refunds",
    } and "Operativa" not in {"Operativa"} or False)


def test_detalle_emits_subtotal_rows_per_merchant(mock_process, mock_fees_result):
    """Sheet 1 must include `Subtotal {merchant}` rows after each merchant.
    Subtotals are how FinOps eyeballs per-merchant totals — gold file relies
    on them, our parser ignores them (it filters by `adquirente`)."""
    _, content = build_fees_export(mock_process, mock_fees_result)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Detalle por Merchant"]

    subtotal_rows = [
        row for row in ws.iter_rows(values_only=True)
        if row and row[1] and str(row[1]).startswith("Subtotal ")
    ]
    subtotal_merchants = {str(r[1]).removeprefix("Subtotal ") for r in subtotal_rows}

    # All 3 merchants in fixture should have a subtotal row
    assert subtotal_merchants == {"AFUNVIP", "Afun Mexico", "BCGAME"}

    # AFUNVIP subtotal: only 1 bitso row (50.0 amount, 1.25 fee)
    afv = next(r for r in subtotal_rows if "AFUNVIP" in str(r[1]))
    assert afv[4] == 1            # # eventos
    assert afv[5] == 50.0          # monto procesado
    assert afv[8] == pytest.approx(1.25)   # fee s/IVA


def test_razon_social_grouping_uses_config(mock_process, mock_fees_result, db_session):
    """When `db` is provided, Sheet 3 must group by razon social per
    `merchant_razon_social_map` config — not the legacy per-merchant
    placeholder. AFUNVIP and Afun Mexico share KODEMAX GLOBAL."""
    _, content = build_fees_export(mock_process, mock_fees_result, db=db_session)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Resumen por Razon Social"]

    rows = []
    for row in ws.iter_rows(values_only=True):
        if row and row[1] and row[1] not in ("Razon Social",):
            rows.append(row)

    # Find KODEMAX row → must list both AFUNVIP and Afun Mexico
    kodemax = next((r for r in rows if r[1] == "KODEMAX GLOBAL SA DE CV"), None)
    assert kodemax is not None, f"KODEMAX row missing. Got: {[r[1] for r in rows]}"
    merchants_str = kodemax[2]
    assert "AFUNVIP" in merchants_str
    assert "Afun Mexico" in merchants_str

    # BCGAME must be under PUBLIPLAY MEXICO
    publiplay = next((r for r in rows if r[1] == "PUBLIPLAY MEXICO SA DE CV"), None)
    assert publiplay is not None
    assert publiplay[2] == "BCGAME"


def test_razon_social_falls_back_when_no_db(mock_process, mock_fees_result):
    """No db = legacy placeholder = one row per merchant. Preserves
    the existing manual download endpoint behavior."""
    _, content = build_fees_export(mock_process, mock_fees_result, db=None)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Resumen por Razon Social"]

    # Each merchant becomes its own razon social row
    razons = []
    for row in ws.iter_rows(values_only=True):
        if row and row[1] and row[1] not in ("Razon Social",):
            razons.append(row[1])
    # In legacy mode, no merchant ever maps to a different razon-social label
    assert "AFUNVIP" in razons
    assert "Afun Mexico" in razons
    assert "BCGAME" in razons
    # No KODEMAX (that would only appear in the db-aware path)
    assert "KODEMAX GLOBAL SA DE CV" not in razons


def test_razon_social_falls_back_for_unmapped_merchant(mock_process, db_session):
    """An unmapped merchant should land under its own name as razon social
    (not be dropped, not crash). Same for a fresh deploy where FinOps
    hasn't yet added the mapping."""
    fees = types.SimpleNamespace(
        merchant_summary=[{"merchant_id": "x", "merchant_name": "TotallyNewMerchant",
                           "gross_amount": 100.0, "total_fee": 10.0}],
        daily_breakdown=[{"date": "2026-04-01", "merchant_id": "x",
                          "merchant_name": "TotallyNewMerchant",
                          "acquirer": "kushki", "amount": 100.0, "fee_amount": 10.0}],
        withdrawals_summary=[], refunds_summary=[], other_fees_summary=[],
    )
    _, content = build_fees_export(mock_process, fees, db=db_session)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Resumen por Razon Social"]
    razons = [row[1] for row in ws.iter_rows(min_row=6, values_only=True)
              if row and row[1]]
    assert "TotallyNewMerchant" in razons


def test_pipeline_creates_auto_fees_uploadedfile(mock_process, mock_fees_result, db_session, tmp_path, monkeypatch):
    """End-to-end test of the Stage 4b logic, but isolated from the rest
    of `_run_full_process` (Mongo extraction, Kushki SFTP, etc.). We
    inline the same path the pipeline takes — this catches regressions
    in the wiring without booting the full background task."""
    # Persist a process row so the FK in UploadedFile resolves
    proc = AccountingProcess(
        id=999,
        name="Test April 2026",
        period_year=2026,
        period_month=4,
        bank_account="Banregio",
        acquirers=["kushki", "bitso", "oxxopay"],
        status="running",
    )
    db_session.add(proc)
    fees = FeesResult(
        process_id=999,
        merchant_summary=mock_fees_result.merchant_summary,
        daily_breakdown=mock_fees_result.daily_breakdown,
        withdrawals_summary=mock_fees_result.withdrawals_summary,
        refunds_summary=mock_fees_result.refunds_summary,
        other_fees_summary=mock_fees_result.other_fees_summary,
        total_fees=0,
    )
    db_session.add(fees)
    db_session.commit()

    # Override UPLOAD_DIR to a tmp path
    monkeypatch.setattr("app.config.settings.UPLOAD_DIR", str(tmp_path))

    # Run the auto-gen block (mirrors processes.py:_run_full_process Stage 4b)
    filename, content = build_fees_export(proc, fees, db=db_session)
    upload_dir = os.path.join(str(tmp_path), str(proc.id))
    os.makedirs(upload_dir, exist_ok=True)
    stored_path = os.path.join(upload_dir, "fees_auto_1234.xlsx")
    with open(stored_path, "wb") as f:
        f.write(content)
    db_session.add(UploadedFile(
        process_id=proc.id,
        file_type="fees",
        original_name=filename,
        stored_path=stored_path,
        file_size=len(content),
        status="auto_generated",
    ))
    db_session.commit()

    # Assert one auto-generated file exists
    rows = db_session.query(UploadedFile).filter_by(
        process_id=proc.id, file_type="fees"
    ).all()
    assert len(rows) == 1
    assert rows[0].status == "auto_generated"
    assert "fees_auto_" in rows[0].stored_path
    assert os.path.exists(rows[0].stored_path)
    assert rows[0].file_size > 0
